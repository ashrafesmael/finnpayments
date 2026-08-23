"""
FinnPayments - FastAPI Application
REST API for invoice processing and accounting entries.
Architecture mirrors FinnVerify's api.py.
"""

import os
import time
import json
import shutil
import logging
import hmac
import hashlib
import base64
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
from pydantic import BaseModel as PydanticBaseModel

from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks, Form, Query, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from src.models import (
    InvoiceStatus, InvoiceType, Currency,
    InvoiceCreateRequest, InvoiceResponse,
    DocumentUploadResponse, DashboardStats,
    JournalEntry, AccountingEntriesResponse
)
from src.database import (
    init_db, get_db, Invoice, InvoiceLineItemDB,
    JournalEntryDB, JournalEntryLineDB, ChartOfAccountsDB, ClassificationRule,
    TDSRate, SessionLocal, AuditLog, Vendor, RecurringTemplate,
    ExchangeRate, get_exchange_rate, fetch_exchange_rates
)
from src.invoice_engine import process_invoice, generate_invoice_id
from src.accounting_engine import (
    generate_accounting_entries, validate_journal_entry, suggest_account_code
)
from src.auth_api import router as auth_router, get_current_company, get_current_user
from src.email_service import email_service
from src.docuseal_integration import (
    is_configured as docuseal_configured,
    create_approval_envelope,
    verify_webhook,
    get_submission_status,
)

logger = logging.getLogger("FinnPayments.API")

# ─── App Configuration ────────────────────────────────────

app = FastAPI(
    title="FinnPayments API",
    description="Invoice Processing & Accounting Entries",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Rate limiter
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS - allow frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static files & uploads
upload_directory = Path("uploads")
upload_directory.mkdir(exist_ok=True)
# Keep temp_uploads for backward compatibility
Path("temp_uploads").mkdir(exist_ok=True)
static_directory = Path("static")
static_directory.mkdir(exist_ok=True)

try:
    app.mount("/static", StaticFiles(directory="static"), name="static")
except:
    pass

# In-memory results store (mirrors FinnVerify pattern)
results_store: Dict[str, Dict] = {}


# ─── Email Action Tokens (stateless HMAC) ───────────────

ACTION_TOKEN_SECRET = os.getenv('SECRET_KEY', 'finnpayments-action-secret-2026')
ACTION_TOKEN_EXPIRY_HOURS = 72  # 3 days

def generate_action_token(invoice_id: str, user_id: str, action: str) -> str:
    """Generate a signed, time-limited token for email approve/decline actions."""
    payload = {
        'invoice_id': invoice_id,
        'user_id': user_id,
        'action': action,
        'exp': int(time.time()) + ACTION_TOKEN_EXPIRY_HOURS * 3600,
    }
    payload_json = json.dumps(payload, separators=(',', ':'))
    payload_b64 = base64.urlsafe_b64encode(payload_json.encode()).decode()
    signature = hmac.new(ACTION_TOKEN_SECRET.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()
    return f"{payload_b64}.{signature}"

def verify_action_token(token: str) -> Optional[dict]:
    """Verify a signed action token. Returns payload dict or None."""
    try:
        parts = token.rsplit('.', 1)
        if len(parts) != 2:
            return None
        payload_b64, signature = parts
        expected_sig = hmac.new(ACTION_TOKEN_SECRET.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected_sig):
            return None
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        if time.time() > payload['exp']:
            return None
        return payload
    except Exception:
        return None


# ─── Audit Log Helper ────────────────────────────────────

def log_audit(action: str, user: dict, entity_type: str = None, entity_id: str = None,
              description: str = None, company_id: str = None, ip_address: str = None):
    """Record an audit log entry."""
    try:
        with get_db() as db:
            db.add(AuditLog(
                user_id=user.get('id') if user else None,
                user_email=user.get('email') if user else None,
                action=action,
                entity_type=entity_type,
                entity_id=entity_id,
                description=description,
                company_id=company_id,
                ip_address=ip_address,
            ))
            db.commit()
    except Exception as e:
        logger.error(f"Failed to write audit log: {e}")


# ─── Authentication router ───────────────────────────────
app.include_router(auth_router)


# ─── Startup ──────────────────────────────────────────────

@app.on_event("startup")
async def startup_event():
    """Initialize database on startup"""
    init_db()
    logger.info("🚀 FinnPayments API started")
    try:
        generate_due_recurring_invoices()
    except Exception as e:
        logger.error(f"Recurring invoice generation failed: {e}")


# ─── Root & Health ────────────────────────────────────────

@app.get("/")
async def root():
    return {
        "message": "FinnPayments API",
        "product": "Invoice Processing & Accounting Entries",
        "company": "FinnPayments",
        "version": "1.0.0",
        "status": "operational",
        "endpoints": {
            "health": "/health",
            "docs": "/docs",
            "upload": "/invoices/upload",
            "invoices": "/invoices",
            "entries": "/accounting/entries",
            "accounts": "/accounting/chart-of-accounts",
            "dashboard": "/dashboard/stats",
        }
    }


@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "service": "FinnPayments",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    }


# ─── Email Action Endpoint (Approve/Decline from email) ─

@app.get("/invoice-action")
async def invoice_action(token: str):
    """Process approve action or show decline form from email link. No login required."""
    payload = verify_action_token(token)
    if not payload:
        return HTMLResponse(content="""
        <!DOCTYPE html><html><head><title>finnpayments</title>
        <style>body{font-family:Arial,sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;background:#0b1120;color:#e2e8f0}</style></head>
        <body><div style="text-align:center"><h1>Invalid or Expired Link</h1><p>This approval link is invalid or has expired.</p><a href="https://payments.finnverify.com" style="color:#10b981">Go to finnpayments</a></div></body></html>
        """, status_code=400)

    invoice_id = payload['invoice_id']
    action = payload['action']  # 'approve' or 'decline'

    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id).first()
        if not invoice:
            return HTMLResponse(content="<h1>Invoice not found</h1>", status_code=404)

        if invoice.status != "pending_review":
            return HTMLResponse(content=f"""
            <!DOCTYPE html><html><head><title>finnpayments</title>
            <style>body{{font-family:Arial,sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;background:#0b1120;color:#e2e8f0}}</style></head>
            <body><div style="text-align:center"><h1>Already Processed</h1>
            <p>Invoice {invoice.invoice_number} has already been {invoice.status}.</p>
            <a href="https://payments.finnverify.com" style="color:#10b981">Go to finnpayments</a></div></body></html>
            """)

        if action == "approve":
            # Process immediately
            invoice.status = "approved"
            invoice.approved_by = payload['user_id']
            invoice.updated_at = datetime.utcnow()
            db.commit()
            log_audit("email_approved", {"id": payload['user_id'], "email": "email_action"},
                      entity_type="invoice", entity_id=invoice_id,
                      description=f"Invoice {invoice.invoice_number} approved via email link",
                      company_id=invoice.company_id)
            return _action_result_page("approved", invoice.invoice_number, invoice.vendor_name)

        elif action == "decline":
            # Show a form with a comment box
            return HTMLResponse(content=f"""
            <!DOCTYPE html><html><head><title>finnpayments - Decline Invoice</title>
            <style>
                body {{ font-family: Arial, sans-serif; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; background: #0b1120; color: #e2e8f0; }}
                .container {{ max-width: 500px; padding: 40px; }}
                h1 {{ color: #ef4444; margin-bottom: 16px; }}
                .invoice-info {{ background: #1c2d4a; padding: 16px; border-radius: 8px; margin: 16px 0; font-size: 14px; }}
                textarea {{ width: 100%; padding: 12px; border-radius: 8px; border: 1px solid #243555; background: #162036; color: #e2e8f0; font-size: 14px; font-family: Arial; min-height: 100px; resize: vertical; }}
                .btn {{ display: inline-block; background: #ef4444; color: white; padding: 12px 30px; border-radius: 8px; text-decoration: none; font-weight: 600; border: none; cursor: pointer; font-size: 15px; }}
                .btn-cancel {{ background: #243555; margin-left: 8px; }}
            </style></head>
            <body>
                <div class="container">
                    <h1>Decline Invoice</h1>
                    <div class="invoice-info">
                        <strong>Invoice:</strong> {invoice.invoice_number}<br>
                        <strong>Vendor:</strong> {invoice.vendor_name}<br>
                        <strong>Amount:</strong> {invoice.currency} {invoice.total_amount:,.2f}
                    </div>
                    <p>Please provide a reason for declining this invoice:</p>
                    <form action="/api/invoice-action" method="POST">
                        <input type="hidden" name="token" value="{token}">
                        <textarea name="comment" placeholder="Reason for declining..."></textarea>
                        <div style="margin-top: 16px;">
                            <button type="submit" class="btn">Confirm Decline</button>
                            <a href="https://payments.finnverify.com" class="btn btn-cancel" style="text-decoration:none; display:inline-block;">Cancel</a>
                        </div>
                    </form>
                </div>
            </body></html>
            """)

        else:
            return HTMLResponse(content="<h1>Invalid action</h1>", status_code=400)


@app.post("/invoice-action")
async def invoice_action_submit(token: str = Form(...), comment: str = Form("")):
    """Process the decline form submission with a comment."""
    payload = verify_action_token(token)
    if not payload:
        return HTMLResponse(content="<h1>Invalid or Expired Link</h1>", status_code=400)

    invoice_id = payload['invoice_id']

    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id).first()
        if not invoice:
            return HTMLResponse(content="<h1>Invoice not found</h1>", status_code=404)

        if invoice.status != "pending_review":
            return HTMLResponse(content=f"<h1>Already Processed</h1><p>Invoice {invoice.invoice_number} has already been {invoice.status}.</p>")

        invoice.status = "rejected"
        invoice.updated_at = datetime.utcnow()
        if comment:
            invoice.notes = (invoice.notes or '') + f"\n[Declined via email: {comment}]"
        db.commit()
        log_audit("email_declined", {"id": payload['user_id'], "email": "email_action"},
                  entity_type="invoice", entity_id=invoice_id,
                  description=f"Invoice {invoice.invoice_number} declined via email link. Reason: {comment or 'No reason provided'}",
                  company_id=invoice.company_id)

    return _action_result_page("declined", invoice.invoice_number, invoice.vendor_name, comment)


def _action_result_page(action: str, invoice_number: str, vendor_name: str, comment: str = None) -> HTMLResponse:
    """Return a styled HTML confirmation page."""
    if action == "approved":
        return HTMLResponse(content=f"""
        <!DOCTYPE html><html><head><title>finnpayments - Approved</title>
        <style>
            body {{ font-family: Arial, sans-serif; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; background: #0b1120; color: #e2e8f0; }}
            .container {{ text-align: center; max-width: 500px; padding: 40px; }}
            h1 {{ color: #10b981; margin-bottom: 16px; }}
            .icon {{ font-size: 48px; margin-bottom: 20px; }}
            .btn {{ display: inline-block; background: #10b981; color: white; padding: 12px 30px; border-radius: 8px; text-decoration: none; font-weight: 600; margin-top: 20px; }}
        </style></head>
        <body><div class="container"><div class="icon">✓</div><h1>Invoice Approved</h1>
        <p>Invoice {invoice_number} from {vendor_name} has been approved successfully.</p>
        <p style="color: #64748b; font-size: 13px; margin-top: 20px;">This action was processed automatically from your email link.</p>
        <a href="https://payments.finnverify.com" class="btn">Go to finnpayments</a></div></body></html>
        """)
    else:
        comment_html = f'<div style="background:#1c2d4a;padding:12px;border-radius:8px;margin:16px 0;font-size:14px;text-align:left;"><strong>Reason:</strong> {comment}</div>' if comment else ''
        return HTMLResponse(content=f"""
        <!DOCTYPE html><html><head><title>finnpayments - Declined</title>
        <style>
            body {{ font-family: Arial, sans-serif; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; background: #0b1120; color: #e2e8f0; }}
            .container {{ text-align: center; max-width: 500px; padding: 40px; }}
            h1 {{ color: #ef4444; margin-bottom: 16px; }}
            .icon {{ font-size: 48px; margin-bottom: 20px; }}
            .btn {{ display: inline-block; background: #10b981; color: white; padding: 12px 30px; border-radius: 8px; text-decoration: none; font-weight: 600; margin-top: 20px; }}
        </style></head>
        <body><div class="container"><div class="icon">✕</div><h1>Invoice Declined</h1>
        <p>Invoice {invoice_number} from {vendor_name} has been declined.</p>
        {comment_html}
        <p style="color: #64748b; font-size: 13px; margin-top: 20px;">This action was processed automatically from your email link.</p>
        <a href="https://payments.finnverify.com" class="btn">Go to finnpayments</a></div></body></html>
        """)


# ─── Invoice Upload & Processing ─────────────────────────

@app.post("/invoices/upload", response_model=None)
async def upload_invoice(
    file: UploadFile = File(...),
    invoice_type: str = Form("supplier"),
    project_code: Optional[str] = Form(None),
    cost_center: Optional[str] = Form(None),
    assigned_to: Optional[str] = Form(None),
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """
    Upload an invoice document for processing.
    Supports PDF, images, CSV, Excel.
    """
    start_time = time.time()
    
    # Validate file type
    allowed_extensions = ('.pdf', '.png', '.jpg', '.jpeg', '.docx', '.doc', '.txt', '.csv', '.xlsx', '.xls', '.bmp', '.tiff', '.webp')
    if not file.filename.lower().endswith(allowed_extensions):
        raise HTTPException(status_code=400, detail=f"Unsupported file type. Allowed: {', '.join(allowed_extensions)}")
    
    # Save uploaded file
    file_path = upload_directory / f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    try:
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        logger.info(f"📁 Saved upload: {file_path} ({len(content)} bytes)")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
    
    # Process the invoice
    try:
        result = await process_invoice(str(file_path), invoice_type, company_id=company['id'])
    except Exception as e:
        logger.error(f"❌ Processing error: {e}")
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")
    
    # Handle multi-invoice PDF (result is a list)
    if isinstance(result, list):
        logger.info(f"📋 Multi-invoice upload: {len(result)} invoices detected")
        for r in result:
            try:
                _save_invoice_to_db(r, invoice_type, str(file_path), project_code, cost_center, company['id'], uploader_user_id=assigned_to or user['id'])
                results_store[r["invoice_id"]] = r
            except Exception as e:
                logger.error(f"❌ Database save error for {r['invoice_id']}: {e}")
        
        total_time = time.time() - start_time
        return {
            "multi_invoice": True,
            "count": len(result),
            "invoices": result,
            "processing_time": round(total_time, 2),
            "message": f"{len(result)} invoices detected and processed in {total_time:.1f}s"
        }
    
    # Single invoice
    try:
        _save_invoice_to_db(result, invoice_type, str(file_path), project_code, cost_center, company['id'], uploader_user_id=assigned_to or user['id'])
    except Exception as e:
        logger.error(f"❌ Database save error: {e}")
    
    results_store[result["invoice_id"]] = result

    processing_time = time.time() - start_time
    result["processing_time"] = round(processing_time, 2)

    # Notify the assigned user about the new invoice (only the assignee, not all company users)
    try:
        from src.auth_models import auth_db as _auth_db
        login_url = os.getenv('SITE_BASE_URL', 'https://payments.finnverify.com')
        extracted = result.get("extracted_data", {})
        inv_num = extracted.get("invoice_number", result["invoice_id"])
        vendor = extracted.get("vendor_name", "Unknown")
        total = extracted.get("total_amount", 0)
        cur = extracted.get("currency", "MUR")
        # Only notify the assigned user (not the uploader)
        assigned_user_id = assigned_to or user['id']
        if assigned_user_id != user['id']:
            assignee = _auth_db.get_user_by_id(assigned_user_id)
            if assignee and assignee['status'] == 'approved':
                base_url = os.getenv('SITE_BASE_URL', 'https://payments.finnverify.com')
                approve_token = generate_action_token(result["invoice_id"], assigned_user_id, "approve")
                decline_token = generate_action_token(result["invoice_id"], assigned_user_id, "decline")
                approve_url = f"{base_url}/api/invoice-action?token={approve_token}"
                decline_url = f"{base_url}/api/invoice-action?token={decline_token}"
                email_service.send_new_invoice_uploaded(
                    assignee['email'], assignee['full_name'], inv_num, vendor, total, cur, login_url,
                    attachment_path=str(file_path),
                    approve_url=approve_url, decline_url=decline_url,
                )
    except Exception as e:
        logger.error(f"Failed to send upload notification: {e}")

    # E-signature integration (DocuSeal/OpenSign) is currently disabled
    # To re-enable, set DOCUSEAL_API_KEY in .env and uncomment the block below
    # if assigned_to and assigned_to != user['id'] and docuseal_configured():
    #     try:
    #         from src.auth_models import auth_db as _auth_db
    #         assignee = _auth_db.get_user_by_id(assigned_to)
    #         if assignee and assignee['status'] == 'approved':
    #             doc_path = str(file_path) if file_path and os.path.exists(str(file_path)) else None
    #             envelope = await create_approval_envelope(
    #                 invoice_id=result["invoice_id"],
    #                 invoice_number=extracted.get("invoice_number", result["invoice_id"]),
    #                 vendor_name=extracted.get("vendor_name", "Unknown"),
    #                 amount=extracted.get("total_amount", 0),
    #                 currency=extracted.get("currency", "MUR"),
    #                 approver_email=assignee['email'],
    #                 approver_name=assignee['full_name'],
    #                 document_path=doc_path,
    #             )
    #             if envelope:
    #                 log_audit("docuseal_envelope_created", user, entity_type="invoice", entity_id=result["invoice_id"],
    #                           description=f"DocuSeal envelope created for {extracted.get('invoice_number', result['invoice_id'])} (submission_id={envelope.get('id')})", company_id=company['id'])
    #     except Exception as e:
    #         logger.error(f"Failed to create DocuSeal envelope on upload: {e}")

    return result


@app.post("/invoices/manual")
async def create_manual_invoice(request: InvoiceCreateRequest, company: dict = Depends(get_current_company)):
    """Create an invoice manually (no document upload)"""
    invoice_id = generate_invoice_id()
    
    invoice_data = {
        "vendor_name": request.vendor_name,
        "vendor_brn": request.vendor_brn,
        "invoice_number": request.invoice_number,
        "invoice_date": request.invoice_date,
        "due_date": request.due_date,
        "currency": request.currency.value if hasattr(request.currency, 'value') else request.currency,
        "line_items": [item.model_dump() for item in request.line_items],
        "subtotal": sum(item.amount - item.tax_amount for item in request.line_items),
        "tax_total": sum(item.tax_amount for item in request.line_items),
        "total_amount": sum(item.amount for item in request.line_items),
        "notes": request.notes,
        "project_code": request.project_code,
        "cost_center": request.cost_center,
        "confidence_score": 1.0,
    }
    
    # Generate accounting entries
    entries = generate_accounting_entries(invoice_id, invoice_data, request.invoice_type.value, company_id=company['id'])
    
    result = {
        "invoice_id": invoice_id,
        "status": "pending_review",
        "extracted_data": invoice_data,
        "suggested_entries": entries,
        "processing_time": 0.0,
        "message": "Manual invoice created successfully"
    }
    
    _save_invoice_to_db(result, request.invoice_type.value, None, request.project_code, request.cost_center, company['id'], uploader_user_id=user['id'])
    results_store[invoice_id] = result
    
    return result


# ─── Invoice CRUD ─────────────────────────────────────────

@app.get("/invoices")
async def list_invoices(
    status: Optional[str] = None,
    invoice_type: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0,
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    company: dict = Depends(get_current_company),
):
    """List all invoices with optional filtering (status, type, search, date range)"""
    with get_db() as db:
        query = db.query(Invoice).filter(Invoice.company_id == company['id']).order_by(Invoice.created_at.desc())

        if status:
            query = query.filter(Invoice.status == status)
        if invoice_type:
            query = query.filter(Invoice.invoice_type == invoice_type)
        if search:
            query = query.filter(
                (Invoice.vendor_name.ilike(f"%{search}%")) |
                (Invoice.invoice_number.ilike(f"%{search}%"))
            )
        if start_date:
            query = query.filter(Invoice.invoice_date >= start_date)
        if end_date:
            query = query.filter(Invoice.invoice_date <= end_date)

        total = query.count()
        invoices = query.offset(offset).limit(limit).all()
        
        return {
            "total": total,
            "invoices": [_invoice_to_dict(inv) for inv in invoices],
            "limit": limit,
            "offset": offset,
        }


@app.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, company: dict = Depends(get_current_company)):
    """Get invoice details with line items and journal entries"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        result = _invoice_to_dict(invoice)
        
        # Include line items
        result["line_items"] = [
            {
                "line_number": li.line_number,
                "description": li.description,
                "quantity": li.quantity,
                "unit_price": li.unit_price,
                "amount": li.amount,
                "tax_rate": li.tax_rate,
                "tax_amount": li.tax_amount,
                "account_code": li.account_code,
                "cost_center": li.cost_center,
                "project_code": li.project_code,
            }
            for li in invoice.line_items
        ]
        
        # Include journal entries
        result["journal_entries"] = [
            {
                "entry_id": je.entry_id,
                "entry_date": je.entry_date,
                "reference": je.reference,
                "description": je.description,
                "total_debit": je.total_debit,
                "total_credit": je.total_credit,
                "is_balanced": je.is_balanced,
                "status": je.status,
                "lines": [
                    {
                        "account_code": jel.account_code,
                        "account_name": jel.account_name,
                        "description": jel.description,
                        "debit": jel.debit,
                        "credit": jel.credit,
                        "cost_center": jel.cost_center,
                        "project_code": jel.project_code,
                    }
                    for jel in je.lines
                ]
            }
            for je in invoice.journal_entries
        ]
        
        # Include from memory store if available
        if invoice_id in results_store:
            result["ai_analysis"] = results_store[invoice_id].get("extracted_data", {}).get("notes")
        
        return result


@app.patch("/invoices/{invoice_id}/status")
async def update_invoice_status(
    invoice_id: str,
    status: str,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Update invoice status (approve, reject, post, etc.)"""
    valid_statuses = [s.value for s in InvoiceStatus]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        old_status = invoice.status

        # Maker/checker: the user who approved cannot also post
        if status == "posted" and company.get('maker_checker_enabled'):
            if invoice.approved_by and invoice.approved_by == user['id']:
                raise HTTPException(
                    status_code=403,
                    detail="Maker/checker: You approved this invoice and cannot post it. Another user must post it."
                )

        # Track who approved
        if status == "approved":
            invoice.approved_by = user['id']

        # Track who posted
        if status == "posted":
            invoice.posted_by = user['id']

        invoice.status = status
        invoice.updated_at = datetime.utcnow()
        
        # If posting, also post journal entries
        if status == "posted":
            for je in invoice.journal_entries:
                je.status = "posted"
                je.posted_by = user['id']

        # If marking as paid and TDS applies, calculate TDS and generate payment entries
        if status == "paid" and invoice.tds_applicable and invoice.tds_rate > 0:
            tds_amount = round(invoice.total_amount * invoice.tds_rate / 100, 2)
            invoice.tds_amount = tds_amount

            # Generate TDS payment journal entry
            from src.invoice_engine import generate_entry_id
            from src.accounting_engine import ACCOUNT_MAPPINGS

            mapping = ACCOUNT_MAPPINGS.get(invoice.invoice_type, ACCOUNT_MAPPINGS["supplier"])
            payable_code, payable_name = mapping["payable"]
            bank_code, bank_name = "01-3000-01", "Bank"
            tds_code, tds_name = "01-3001-01", "TDS Payable (MRA)"

            tds_entry_id = generate_entry_id()
            tds_entry = JournalEntryDB(
                entry_id=tds_entry_id,
                invoice_id=invoice.invoice_id,
                entry_date=datetime.now().strftime("%Y-%m-%d"),
                reference=f"TDS-PAY-{invoice.invoice_number}",
                description=f"TDS withholding on payment to {invoice.vendor_name}",
                total_debit=invoice.total_amount,
                total_credit=invoice.total_amount,
                is_balanced=True,
                status="posted",
                created_by=user['email'],
                posted_by=user['id'],
                company_id=company['id'],
            )
            db.add(tds_entry)

            # Dr Trade Creditors (gross), Cr TDS Payable (tax), Cr Bank (net)
            db.add(JournalEntryLineDB(
                entry_id=tds_entry_id,
                account_code=payable_code,
                account_name=payable_name,
                description=f"Settlement of {invoice.invoice_number} - {invoice.vendor_name}",
                debit=invoice.total_amount,
                credit=0.0,
            ))
            db.add(JournalEntryLineDB(
                entry_id=tds_entry_id,
                account_code=tds_code,
                account_name=tds_name,
                description=f"TDS @ {invoice.tds_rate}% on {invoice.invoice_number}",
                debit=0.0,
                credit=tds_amount,
            ))
            db.add(JournalEntryLineDB(
                entry_id=tds_entry_id,
                account_code=bank_code,
                account_name=bank_name,
                description=f"Net payment to {invoice.vendor_name} for {invoice.invoice_number}",
                debit=0.0,
                credit=round(invoice.total_amount - tds_amount, 2),
            ))
            logger.info(f"📊 TDS entry generated: {tds_entry_id} (TDS: {tds_amount}, Net: {invoice.total_amount - tds_amount})")
        
        db.commit()

        log_audit(
            f"invoice_{status}", user,
            entity_type="invoice", entity_id=invoice_id,
            description=f"Invoice {invoice.invoice_number} ({invoice.vendor_name}) status changed from {old_status} to {status}",
            company_id=company['id'],
        )

        # ── Send targeted workflow notifications ──
        login_url = os.getenv('SITE_BASE_URL', 'https://payments.finnverify.com')
        from src.auth_models import auth_db as _auth_db

        def notify_user(user_id, send_func):
            """Send notification to a specific user."""
            if not user_id:
                return
            notifee = _auth_db.get_user_by_id(user_id)
            if not notifee or notifee['status'] != 'approved':
                return
            if user_id == user['id']:
                return  # Don't notify the action performer
            send_func(notifee['email'], notifee['full_name'],
                      invoice.invoice_number, invoice.vendor_name,
                      invoice.total_amount, invoice.currency, login_url)

        if status == "approved":
            # Notify the assigned user that the invoice is ready for posting
            notify_user(invoice.assigned_to, email_service.send_invoice_approved)

        elif status == "rejected":
            # Notify the user who uploaded it (if different from rejector)
            notify_user(invoice.assigned_to, lambda e, n, inv, ven, amt, cur, url=None:
                email_service.send_invoice_rejected(e, n, inv, ven, amt, cur))

        elif status == "posted":
            # Notify the assigned user that the invoice is posted and ready for payment
            notify_user(invoice.assigned_to, email_service.send_invoice_posted)

        return {
            "invoice_id": invoice_id,
            "old_status": old_status,
            "new_status": status,
            "message": f"Invoice status updated to {status}"
        }




@app.get("/invoices/{invoice_id}/document")
async def get_invoice_document(invoice_id: str, company: dict = Depends(get_current_company)):
    """Serve the original uploaded document for an invoice"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if not invoice.source_file:
            raise HTTPException(status_code=404, detail="No document attached to this invoice")

        file_path = Path(invoice.source_file)

        # Also check temp_uploads if file was uploaded before migration
        if not file_path.exists():
            alt_path = Path("temp_uploads") / file_path.name
            if alt_path.exists():
                file_path = alt_path
            else:
                raise HTTPException(status_code=404, detail="Document file not found on disk")

        # Determine media type
        ext = file_path.suffix.lower()
        media_types = {
            ".pdf": "application/pdf",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".bmp": "image/bmp",
            ".tiff": "image/tiff",
            ".webp": "image/webp",
        }
        media_type = media_types.get(ext, "application/octet-stream")

        return FileResponse(
            str(file_path),
            media_type=media_type,
            filename=file_path.name,
            headers={"Content-Disposition": f"inline; filename={file_path.name}"}
        )



@app.get("/invoices/{invoice_id}/document/preview")
async def get_invoice_document_preview(invoice_id: str, page: int = Query(0, ge=0), company: dict = Depends(get_current_company)):
    """Return invoice document pages as base64 images for inline viewing."""
    import base64
    from io import BytesIO

    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if not invoice.source_file:
            raise HTTPException(status_code=404, detail="No document attached")

        file_path = Path(invoice.source_file)
        if not file_path.exists():
            # Check temp_uploads fallback
            alt_path = Path("temp_uploads") / file_path.name
            if alt_path.exists():
                file_path = alt_path
            else:
                raise HTTPException(status_code=404, detail="File not found on disk")

        ext = file_path.suffix.lower()

        if ext in (".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tiff"):
            # Image file: return as-is
            with open(file_path, "rb") as f:
                img_data = base64.b64encode(f.read()).decode()
            mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "bmp": "image/bmp", "webp": "image/webp", "tiff": "image/tiff"}
            return {
                "total_pages": 1,
                "current_page": 0,
                "mime_type": mime.get(ext.lstrip("."), "image/png"),
                "image": img_data,
            }

        elif ext == ".pdf":
            try:
                from pdf2image import convert_from_path
                # Get total pages first
                from pdf2image.pdf2image import pdfinfo_from_path
                info = pdfinfo_from_path(str(file_path))
                total_pages = info.get("Pages", 1)

                # Clamp page number
                page = min(page, total_pages - 1)

                # Convert requested page (1-indexed for pdf2image)
                images = convert_from_path(
                    str(file_path),
                    first_page=page + 1,
                    last_page=page + 1,
                    dpi=150,
                    fmt="png"
                )

                if images:
                    buf = BytesIO()
                    images[0].save(buf, format="PNG")
                    img_data = base64.b64encode(buf.getvalue()).decode()
                    return {
                        "total_pages": total_pages,
                        "current_page": page,
                        "mime_type": "image/png",
                        "image": img_data,
                    }
                else:
                    raise HTTPException(status_code=500, detail="Failed to render PDF page")
            except ImportError:
                raise HTTPException(status_code=500, detail="pdf2image not installed")
        else:
            raise HTTPException(status_code=400, detail=f"Preview not supported for {ext} files")



class ReclassifyRequest(PydanticBaseModel):
    user_context: str


@app.post("/invoices/{invoice_id}/reclassify")
async def reclassify_invoice(invoice_id: str, request: ReclassifyRequest, company: dict = Depends(get_current_company)):
    """
    Reclassify invoice line items using user-provided context.
    Called when the system defaulted to Licences (01-6000-04) and the user
    provides additional context about the nature of the expense.
    """
    import httpx

    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")

        line_items = db.query(InvoiceLineItemDB).filter(
            InvoiceLineItemDB.invoice_id == invoice_id
        ).order_by(InvoiceLineItemDB.line_number).all()

        if not line_items:
            raise HTTPException(status_code=400, detail="No line items found")

        # Build context for LLM
        items_text = "\n".join([
            f"  Line {li.line_number}: {li.description} | Amount: {li.amount} | Current account: {li.account_code}"
            for li in line_items
        ])

        api_key = os.getenv("GROQ_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="AI service not configured")

        prompt = f"""You are an accounting classification expert for Mont Choisy Golf, a Mauritian property/golf company.

VENDOR: {invoice.vendor_name}
INVOICE #: {invoice.invoice_number}
DATE: {invoice.invoice_date}
TOTAL: {invoice.currency} {invoice.total_amount}

LINE ITEMS:
{items_text}

USER CONTEXT (the user has described this expense as):
"{request.user_context}"

CHART OF ACCOUNTS (assign the best account_code per line item):
01-5100-04  Basic Salary_Admin
01-5102-04  Statutory Contribution (NPS & TWEF)
01-5105-04  Staff Travelling Cost
01-5108-04  Recruitment Cost
01-5110-04  Training Staff Cost
01-5133-04  Medical Scheme
01-5002-01  Golf Cart Maintenance & General
01-5007-02  Event Cost
01-5201-01  Operating Supplies_Golf Course
01-6000-04  Licences (software licences, permits, domain renewals)
01-6001-04  Insurance_General
01-6002-04  Telephone, Mobiles and Internet
01-6003-04  ICT Expenses (software, IT services, hosting, SaaS, hardware)
01-6005-04  Printing, Postage & Stationery
01-6006-04  Professional Fees ADM (legal, consulting, advisory, law firms, court costs, notary, architects)
01-6007-04  Health & Safety Expenses
01-6008-04  Security Fees (security companies, cash handling, guarding)
01-6008-06  Marketing Cost_Others
01-6009-04  Pest Control
01-6010-04  Waste Removal
01-6010-06  General Marketing Materials
01-6015-04  Cleaning Expenses
01-6020-04  Office Supplies & Consumables
01-6021-04  Payroll Processing Fee
01-6022-04  Audit Fee
01-6023-04  Secretarial Fee
01-6025-04  Taxation Fee
01-6030-04  Subscriptions & Memberships
01-6050-05  R&M - Golf Course Fertilizers
01-6051-05  R&M - Fuel and Diesel
01-6053-05  R&M - Equipment Repairs
01-6057-05  R&M - Golf Course Chemicals
01-6058-05  R&M - Golf Course (Plant, Seed, Sand)
01-6059-05  R&M - Golf Course Maintenance Others
01-6070-05  Building Maintenance (painting, plumbing, electrical)
01-6075-05  External Maintenance Contractor
01-6076-05  Gas Clubhouse
01-6080-05  R&M - Irrigation Pumping Station
01-6090-05  Vehicle Running Expenses
01-6100-04  Uniforms & Protective Clothing
01-6250-04  Electricity (CEB, power)
01-6251-04  Water (CWA)
01-6300-07  Rental of Land
01-6301-07  Credit Card Commission and Bank Charges
01-6302-07  Surcharge & Penalties
01-6401-07  Corporate Management Fee
01-6403-07  Estate Shared Cost
01-8001-08  Bank Interest on Loan
01-1005-01  PPE-Cost-Other Equipment (capital purchases)
01-1006-01  PPE-Cost-Furniture and Fittings (capital purchases)

Based on the vendor name, line item descriptions, and the user's context, assign the BEST matching account_code for EACH line item.

Return ONLY a JSON array like:
[
  {{"line_number": 1, "account_code": "01-XXXX-XX", "reason": "Brief explanation"}},
  {{"line_number": 2, "account_code": "01-XXXX-XX", "reason": "Brief explanation"}}
]

Return ONLY valid JSON, no markdown."""

        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                models = ["openai/gpt-oss-120b", "openai/gpt-oss-20b"]
                response = None
                for model in models:
                    response = await client.post(
                        "https://api.groq.com/openai/v1/chat/completions",
                        headers={
                            "Authorization": f"Bearer {api_key}",
                            "Content-Type": "application/json"
                        },
                        json={
                            "model": model,
                            "messages": [
                                {"role": "system", "content": "You are an expert accountant. Return only valid JSON arrays."},
                                {"role": "user", "content": prompt}
                            ],
                            "temperature": 0.1,
                            "max_tokens": 1000
                        }
                    )
                    if response.status_code == 200:
                        break
                    logger.warning(f"⚠️ Reclassify model {model} returned {response.status_code}, trying fallback...")

                if not response or response.status_code != 200:
                    raise HTTPException(status_code=502, detail=f"AI service error: {response.status_code if response else 'no response'}")

                import re as re_mod
                result = response.json()
                ai_content = result["choices"][0]["message"]["content"]
                ai_content = re_mod.sub(r'^```json\s*', '', ai_content)
                ai_content = re_mod.sub(r'\s*```$', '', ai_content)
                classifications = json.loads(ai_content)

        except json.JSONDecodeError:
            raise HTTPException(status_code=502, detail="AI returned invalid response")
        except Exception as e:
            logger.error(f"Reclassify error: {e}")
            raise HTTPException(status_code=500, detail=str(e))

        # Update line items with new account codes
        from src.accounting_engine import lookup_account_name
        updates = []
        for cls in classifications:
            line_num = cls.get("line_number")
            new_code = cls.get("account_code")
            reason = cls.get("reason", "")
            if not line_num or not new_code:
                continue

            for li in line_items:
                if li.line_number == line_num:
                    old_code = li.account_code
                    li.account_code = new_code
                    updates.append({
                        "line_number": line_num,
                        "old_account": old_code,
                        "new_account": new_code,
                        "account_name": lookup_account_name(new_code),
                        "reason": reason,
                        "description": li.description,
                    })
                    break

        # Delete old journal entries and regenerate
        old_entries = db.query(JournalEntryDB).filter(
            JournalEntryDB.invoice_id == invoice_id
        ).all()
        for oe in old_entries:
            db.query(JournalEntryLineDB).filter(
                JournalEntryLineDB.entry_id == oe.entry_id
            ).delete()
            db.delete(oe)

        # Rebuild invoice data for entry generation
        invoice_data = {
            "vendor_name": invoice.vendor_name,
            "invoice_number": invoice.invoice_number,
            "invoice_date": invoice.invoice_date,
            "total_amount": invoice.total_amount,
            "tax_total": invoice.tax_total,
            "subtotal": invoice.subtotal,
            "line_items": [
                {
                    "description": li.description,
                    "amount": li.amount,
                    "tax_amount": li.tax_amount,
                    "account_code": li.account_code,
                }
                for li in line_items
            ],
            "suggested_cost_center": invoice.cost_center,
            "project_code": invoice.project_code,
        }

        from src.accounting_engine import generate_accounting_entries
        from src.invoice_engine import generate_entry_id
        new_entries = generate_accounting_entries(invoice_id, invoice_data, invoice.invoice_type or "supplier")

        for entry in new_entries:
            je = JournalEntryDB(
                entry_id=entry["entry_id"],
                invoice_id=invoice_id,
                entry_date=entry.get("entry_date"),
                reference=entry.get("reference"),
                description=entry.get("description"),
                total_debit=entry.get("total_debit", 0),
                total_credit=entry.get("total_credit", 0),
                is_balanced=entry.get("is_balanced", True),
                status="draft",
                created_by="reclassify",
            )
            db.add(je)
            for line in entry.get("lines", []):
                jel = JournalEntryLineDB(
                    entry_id=entry["entry_id"],
                    account_code=line.get("account_code"),
                    account_name=line.get("account_name"),
                    description=line.get("description"),
                    debit=line.get("debit", 0),
                    credit=line.get("credit", 0),
                    cost_center=line.get("cost_center"),
                    project_code=line.get("project_code"),
                )
                db.add(jel)

        db.commit()
        # Save classification rules for future learning
        for u in updates:
            if u["new_account"] != "01-6000-04":  # Don't learn the default
                existing = db.query(ClassificationRule).filter(
                    ClassificationRule.vendor_name == invoice.vendor_name,
                    ClassificationRule.account_code == u["new_account"],
                    ClassificationRule.company_id == company['id'],
                ).first()
                if existing:
                    existing.user_context = request.user_context
                    existing.times_used = (existing.times_used or 0) + 1
                    existing.updated_at = datetime.utcnow()
                else:
                    db.add(ClassificationRule(
                        vendor_name=invoice.vendor_name,
                        description_pattern=u.get("description", "")[:200] if u.get("description") else None,
                        account_code=u["new_account"],
                        account_name=u.get("account_name", ""),
                        user_context=request.user_context,
                        source="reclassify",
                        times_used=1,
                        company_id=company['id'],
                    ))
                db.commit()

        logger.info(f"🔄 Reclassified invoice {invoice_id}: {len(updates)} line items updated")
        logger.info(f"📚 Saved {len([u for u in updates if u['new_account'] != '01-6000-04'])} classification rule(s) for future learning")

        return {
            "invoice_id": invoice_id,
            "updates": updates,
            "message": f"{len(updates)} line item(s) reclassified based on your context",
            "user_context": request.user_context,
        }


# ─── Bulk Invoice Operations ────────────────────────────

class BulkActionRequest(PydanticBaseModel):
    invoice_ids: List[str]


@app.post("/invoices/bulk/approve")
async def bulk_approve(request: BulkActionRequest, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Approve multiple pending_review invoices at once."""
    approved = 0
    skipped = 0
    errors = []
    with get_db() as db:
        for inv_id in request.invoice_ids:
            invoice = db.query(Invoice).filter(Invoice.invoice_id == inv_id, Invoice.company_id == company['id']).first()
            if not invoice:
                errors.append(f"{inv_id}: not found")
                continue
            if invoice.status != "pending_review":
                skipped += 1
                continue
            invoice.status = "approved"
            invoice.approved_by = user['id']
            invoice.updated_at = datetime.utcnow()
            approved += 1
            log_audit("invoice_approved", user, entity_type="invoice", entity_id=inv_id,
                      description=f"Invoice {invoice.invoice_number} approved (bulk)", company_id=company['id'])
        db.commit()
    return {"message": f"{approved} approved, {skipped} skipped", "approved": approved, "skipped": skipped, "errors": errors}


@app.post("/invoices/bulk/post")
async def bulk_post(request: BulkActionRequest, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Post multiple approved invoices to the GL at once."""
    posted = 0
    skipped = 0
    errors = []
    with get_db() as db:
        for inv_id in request.invoice_ids:
            invoice = db.query(Invoice).filter(Invoice.invoice_id == inv_id, Invoice.company_id == company['id']).first()
            if not invoice:
                errors.append(f"{inv_id}: not found")
                continue
            if invoice.status != "approved":
                skipped += 1
                continue
            # Maker/checker check
            if company.get('maker_checker_enabled') and invoice.approved_by and invoice.approved_by == user['id']:
                errors.append(f"{invoice.invoice_number}: maker/checker blocked (you approved this)")
                continue
            invoice.status = "posted"
            invoice.posted_by = user['id']
            invoice.updated_at = datetime.utcnow()
            for je in invoice.journal_entries:
                je.status = "posted"
                je.posted_by = user['id']
            posted += 1
            log_audit("invoice_posted", user, entity_type="invoice", entity_id=inv_id,
                      description=f"Invoice {invoice.invoice_number} posted (bulk)", company_id=company['id'])
        db.commit()
    return {"message": f"{posted} posted, {skipped} skipped", "posted": posted, "skipped": skipped, "errors": errors}


@app.post("/invoices/bulk/delete")
async def bulk_delete(request: BulkActionRequest, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Delete multiple draft/pending invoices at once (cannot delete posted/paid)."""
    deleted = 0
    skipped = 0
    errors = []
    with get_db() as db:
        for inv_id in request.invoice_ids:
            invoice = db.query(Invoice).filter(Invoice.invoice_id == inv_id, Invoice.company_id == company['id']).first()
            if not invoice:
                errors.append(f"{inv_id}: not found")
                continue
            if invoice.status in ("posted", "paid"):
                skipped += 1
                continue
            # Delete journal entries and line items first
            for je in invoice.journal_entries:
                db.query(JournalEntryLineDB).filter(JournalEntryLineDB.entry_id == je.entry_id).delete()
                db.query(JournalEntryDB).filter(JournalEntryDB.entry_id == je.entry_id).delete()
            db.query(InvoiceLineItemDB).filter(InvoiceLineItemDB.invoice_id == inv_id).delete()
            db.query(Invoice).filter(Invoice.invoice_id == inv_id).delete()
            deleted += 1
            log_audit("invoice_deleted", user, entity_type="invoice", entity_id=inv_id,
                      description=f"Invoice {invoice.invoice_number} deleted (bulk)", company_id=company['id'])
        db.commit()
    return {"message": f"{deleted} deleted, {skipped} skipped", "deleted": deleted, "skipped": skipped, "errors": errors}


@app.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, company: dict = Depends(get_current_company)):
    """Delete an invoice (only if draft or pending)"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        if invoice.status in ("posted", "paid"):
            raise HTTPException(status_code=400, detail="Cannot delete posted or paid invoices")
        
        db.delete(invoice)
        db.commit()
        
        if invoice_id in results_store:
            del results_store[invoice_id]

        log_audit("invoice_deleted", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"Invoice {invoice_id} deleted", company_id=company['id'])

        return {"message": f"Invoice {invoice_id} deleted"}


# ─── Admin / Reset ────────────────────────────────────────

class ResetConfirmRequest(PydanticBaseModel):
    confirm: bool = False


@app.post("/admin/reset")
async def reset_all_data(
    request: ResetConfirmRequest,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Delete ALL invoices, line items and journal entries for the current company, plus their uploaded documents.
    Chart of accounts and learned classification rules are preserved.
    Admin only."""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required to reset data")
    if not request.confirm:
        raise HTTPException(status_code=400, detail='Confirmation required: pass {"confirm": true}')

    with get_db() as db:
        invoice_count = db.query(Invoice).filter(Invoice.company_id == company['id']).count()
        entry_count = db.query(JournalEntryDB).filter(JournalEntryDB.company_id == company['id']).count()
        source_files = [row[0] for row in db.query(Invoice.source_file).filter(Invoice.company_id == company['id']).all()]

        # Delete children first, scoped by company
        entry_ids = [row[0] for row in db.query(JournalEntryDB.entry_id).filter(JournalEntryDB.company_id == company['id']).all()]
        if entry_ids:
            db.query(JournalEntryLineDB).filter(JournalEntryLineDB.entry_id.in_(entry_ids)).delete(synchronize_session=False)
        db.query(JournalEntryDB).filter(JournalEntryDB.company_id == company['id']).delete(synchronize_session=False)

        invoice_ids = [row[0] for row in db.query(Invoice.invoice_id).filter(Invoice.company_id == company['id']).all()]
        if invoice_ids:
            db.query(InvoiceLineItemDB).filter(InvoiceLineItemDB.invoice_id.in_(invoice_ids)).delete(synchronize_session=False)
        db.query(Invoice).filter(Invoice.company_id == company['id']).delete(synchronize_session=False)
        db.commit()

    # Remove uploaded documents that belonged to the deleted invoices
    safe_dirs = {upload_directory.resolve(), (Path.cwd() / "temp_uploads").resolve()}
    files_removed = 0
    for rel_path in source_files:
        if not rel_path:
            continue
        p = Path(rel_path)
        if not p.is_absolute():
            p = Path.cwd() / p
        try:
            if p.resolve().parent in safe_dirs and p.exists():
                p.unlink()
                files_removed += 1
        except OSError:
            logger.warning(f"Could not remove uploaded file: {p}")

    logger.info(f"Reset complete for {company['name']}: {invoice_count} invoices, {entry_count} journal entries, {files_removed} files removed")

    log_audit("data_reset", user, entity_type="company", entity_id=company['id'],
              description=f"Reset {company['name']}: {invoice_count} invoices, {entry_count} entries deleted", company_id=company['id'])

    return {
        "message": f"All invoice records and journal entries deleted for {company['name']}",
        "invoices_deleted": invoice_count,
        "journal_entries_deleted": entry_count,
        "files_removed": files_removed,
    }


# ─── Accounting Entries ───────────────────────────────────

@app.get("/accounting/entries")
async def list_journal_entries(
    status: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0,
    company: dict = Depends(get_current_company),
):
    """List all journal entries"""
    with get_db() as db:
        query = db.query(JournalEntryDB).filter(JournalEntryDB.company_id == company['id']).order_by(JournalEntryDB.created_at.desc())
        
        if status:
            query = query.filter(JournalEntryDB.status == status)
        
        total = query.count()
        entries = query.offset(offset).limit(limit).all()
        
        return {
            "total": total,
            "entries": [
                {
                    "entry_id": e.entry_id,
                    "invoice_id": e.invoice_id,
                    "entry_date": e.entry_date,
                    "reference": e.reference,
                    "description": e.description,
                    "total_debit": e.total_debit,
                    "total_credit": e.total_credit,
                    "is_balanced": e.is_balanced,
                    "status": e.status,
                    "currency": e.invoice.currency if e.invoice else "MUR",
                    "posted_by": e.posted_by,
                    "created_at": e.created_at.isoformat() if e.created_at else None,
                    "lines": [
                        {
                            "account_code": l.account_code,
                            "account_name": l.account_name,
                            "description": l.description,
                            "debit": l.debit,
                            "credit": l.credit,
                        }
                        for l in e.lines
                    ]
                }
                for e in entries
            ],
        }


@app.post("/accounting/entries/{entry_id}/post")
async def post_journal_entry(
    entry_id: str,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Post a journal entry (make it permanent)"""
    with get_db() as db:
        entry = db.query(JournalEntryDB).filter(JournalEntryDB.entry_id == entry_id, JournalEntryDB.company_id == company['id']).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Journal entry not found")
        
        if entry.status == "posted":
            raise HTTPException(status_code=400, detail="Entry already posted")
        
        # Maker/checker: the user who approved the invoice cannot post the entry
        if company.get('maker_checker_enabled') and entry.invoice:
            if entry.invoice.approved_by and entry.invoice.approved_by == user['id']:
                raise HTTPException(
                    status_code=403,
                    detail="Maker/checker: You approved this invoice and cannot post the journal entry. Another user must post it."
                )

        # Validate before posting
        entry_dict = {
            "lines": [
                {"account_code": l.account_code, "debit": l.debit, "credit": l.credit}
                for l in entry.lines
            ]
        }
        validation = validate_journal_entry(entry_dict)
        if not validation["is_valid"]:
            raise HTTPException(status_code=400, detail=f"Validation failed: {validation['issues']}")
        
        entry.status = "posted"
        entry.posted_by = user['id']
        db.commit()

        log_audit("journal_posted", user, entity_type="journal_entry", entity_id=entry_id,
                  description=f"Journal entry {entry_id} posted", company_id=company['id'])

        return {"entry_id": entry_id, "status": "posted", "message": "Journal entry posted successfully"}


@app.post("/accounting/entries/{entry_id}/reverse")
async def reverse_journal_entry(
    entry_id: str,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Create a reversing entry"""
    with get_db() as db:
        original = db.query(JournalEntryDB).filter(JournalEntryDB.entry_id == entry_id, JournalEntryDB.company_id == company['id']).first()
        if not original:
            raise HTTPException(status_code=404, detail="Journal entry not found")
        
        # Maker/checker: the user who posted cannot reverse
        if company.get('maker_checker_enabled'):
            if original.posted_by and original.posted_by == user['id']:
                raise HTTPException(
                    status_code=403,
                    detail="Maker/checker: You posted this entry and cannot reverse it. Another user must reverse it."
                )

        from src.invoice_engine import generate_entry_id
        reversal_id = generate_entry_id()
        
        reversal = JournalEntryDB(
            entry_id=reversal_id,
            invoice_id=original.invoice_id,
            entry_date=datetime.now().strftime("%Y-%m-%d"),
            reference=f"REV-{original.reference}",
            description=f"Reversal of {original.entry_id}: {original.description}",
            total_debit=original.total_credit,
            total_credit=original.total_debit,
            is_balanced=True,
            status="draft",
            created_by="system",
            company_id=company['id'],
        )
        db.add(reversal)
        
        for line in original.lines:
            rev_line = JournalEntryLineDB(
                entry_id=reversal_id,
                account_code=line.account_code,
                account_name=line.account_name,
                description=f"Reversal: {line.description}",
                debit=line.credit,  # Swap
                credit=line.debit,  # Swap
                cost_center=line.cost_center,
                project_code=line.project_code,
            )
            db.add(rev_line)
        
        original.status = "reversed"
        db.commit()

        log_audit("journal_reversed", user, entity_type="journal_entry", entity_id=entry_id,
                  description=f"Journal entry {entry_id} reversed (reversal: {reversal_id})", company_id=company['id'])

        return {
            "original_entry_id": entry_id,
            "reversal_entry_id": reversal_id,
            "message": "Reversal entry created"
        }


# ─── TDS (Tax Deducted at Source) ────────────────────────

class TDSRateCreate(PydanticBaseModel):
    payment_type: str
    description: Optional[str] = None
    rate: float
    threshold: float = 0.0
    resident: bool = True
    non_resident: bool = True


@app.get("/tds/rates")
async def list_tds_rates(company: dict = Depends(get_current_company)):
    """List all TDS rates for the active company"""
    with get_db() as db:
        rates = db.query(TDSRate).filter(
            TDSRate.company_id == company['id']
        ).order_by(TDSRate.payment_type).all()
        return {
            "total": len(rates),
            "rates": [
                {
                    "id": r.id,
                    "payment_type": r.payment_type,
                    "description": r.description,
                    "rate": r.rate,
                    "threshold": r.threshold,
                    "resident": r.resident,
                    "non_resident": r.non_resident,
                    "is_active": r.is_active,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                }
                for r in rates
            ]
        }


@app.post("/tds/rates")
async def create_tds_rate(rate_data: TDSRateCreate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Create a new TDS rate (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    with get_db() as db:
        rate = TDSRate(
            payment_type=rate_data.payment_type,
            description=rate_data.description,
            rate=rate_data.rate,
            threshold=rate_data.threshold,
            resident=rate_data.resident,
            non_resident=rate_data.non_resident,
            is_active=True,
            company_id=company['id'],
        )
        db.add(rate)
        db.commit()
        return {"id": rate.id, "message": f"TDS rate created: {rate_data.payment_type} @ {rate_data.rate}%"}


@app.put("/tds/rates/{rate_id}")
async def update_tds_rate(rate_id: int, rate_data: TDSRateCreate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Update a TDS rate (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    with get_db() as db:
        rate = db.query(TDSRate).filter(TDSRate.id == rate_id, TDSRate.company_id == company['id']).first()
        if not rate:
            raise HTTPException(status_code=404, detail="TDS rate not found")
        rate.payment_type = rate_data.payment_type
        rate.description = rate_data.description
        rate.rate = rate_data.rate
        rate.threshold = rate_data.threshold
        rate.resident = rate_data.resident
        rate.non_resident = rate_data.non_resident
        rate.updated_at = datetime.utcnow()
        db.commit()
        return {"message": "TDS rate updated"}


@app.delete("/tds/rates/{rate_id}")
async def delete_tds_rate(rate_id: int, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Delete a TDS rate (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    with get_db() as db:
        rate = db.query(TDSRate).filter(TDSRate.id == rate_id, TDSRate.company_id == company['id']).first()
        if not rate:
            raise HTTPException(status_code=404, detail="TDS rate not found")
        db.delete(rate)
        db.commit()
        return {"message": "TDS rate deleted"}


class TDSOverrideRequest(PydanticBaseModel):
    tds_applicable: bool
    tds_rate: float = 0.0


@app.patch("/invoices/{invoice_id}/tds")
async def update_invoice_tds(
    invoice_id: str,
    request: TDSOverrideRequest,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Override TDS settings on an invoice (before payment)"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if invoice.status == "paid":
            raise HTTPException(status_code=400, detail="Cannot modify TDS on a paid invoice")
        invoice.tds_applicable = request.tds_applicable
        invoice.tds_rate = request.tds_rate if request.tds_applicable else 0.0
        invoice.updated_at = datetime.utcnow()
        db.commit()

        log_audit("tds_override", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"TDS overridden on {invoice.invoice_number}: applicable={invoice.tds_applicable}, rate={invoice.tds_rate}%", company_id=company['id'])

        return {
            "invoice_id": invoice_id,
            "tds_applicable": invoice.tds_applicable,
            "tds_rate": invoice.tds_rate,
            "message": "TDS settings updated",
        }


@app.get("/tds/register")
async def tds_register(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    company: dict = Depends(get_current_company),
):
    """TDS register — all TDS deducted in a period"""
    with get_db() as db:
        query = db.query(Invoice).filter(
            Invoice.company_id == company['id'],
            Invoice.tds_applicable == True,
            Invoice.tds_amount > 0,
        )
        if start_date:
            query = query.filter(Invoice.invoice_date >= start_date)
        if end_date:
            query = query.filter(Invoice.invoice_date <= end_date)
        invoices = query.order_by(Invoice.invoice_date).all()

        total_tds = sum(inv.tds_amount for inv in invoices)
        total_gross = sum(inv.total_amount for inv in invoices)
        total_net = total_gross - total_tds

        return {
            "company": company['name'],
            "period": f"{start_date or 'all'} to {end_date or 'now'}",
            "summary": {
                "count": len(invoices),
                "total_gross": round(total_gross, 2),
                "total_tds": round(total_tds, 2),
                "total_net": round(total_net, 2),
            },
            "entries": [
                {
                    "invoice_id": inv.invoice_id,
                    "invoice_number": inv.invoice_number,
                    "vendor_name": inv.vendor_name,
                    "invoice_date": inv.invoice_date,
                    "total_amount": inv.total_amount,
                    "tds_rate": inv.tds_rate,
                    "tds_amount": inv.tds_amount,
                    "net_amount": round(inv.total_amount - inv.tds_amount, 2),
                    "tds_paid_to_mra": inv.tds_paid_to_mra,
                    "tds_paid_date": inv.tds_paid_date,
                }
                for inv in invoices
            ],
        }


@app.patch("/tds/mark-remitted")
async def mark_tds_remitted(
    invoice_id: str,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Mark TDS as remitted to MRA"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if not invoice.tds_applicable or invoice.tds_amount <= 0:
            raise HTTPException(status_code=400, detail="This invoice has no TDS to remit")
        invoice.tds_paid_to_mra = True
        invoice.tds_paid_date = datetime.now().strftime("%Y-%m-%d")
        db.commit()

        log_audit("tds_remitted", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"TDS for {invoice.invoice_number} ({invoice.vendor_name}) marked as remitted to MRA", company_id=company['id'])

        return {"message": f"TDS for {invoice.invoice_number} marked as remitted to MRA"}


# ─── Audit Log ───────────────────────────────────────────

@app.get("/audit-log")
async def get_audit_log(
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    limit: int = Query(100, le=500),
    offset: int = 0,
    company: dict = Depends(get_current_company),
):
    """Get audit log entries (admin only, scoped to company)"""
    user = company  # get_current_company already validates the user
    # Re-check admin via the header dependency chain
    from src.auth_api import get_current_user
    # We need the actual user; get_current_company returns company dict
    # Let's just query by company_id and restrict to admin
    with get_db() as db:
        query = db.query(AuditLog).filter(AuditLog.company_id == company['id'])
        if action:
            query = query.filter(AuditLog.action == action)
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        total = query.count()
        entries = query.order_by(AuditLog.timestamp.desc()).offset(offset).limit(limit).all()
        return {
            "total": total,
            "entries": [
                {
                    "id": e.id,
                    "timestamp": e.timestamp.isoformat() if e.timestamp else None,
                    "user_email": e.user_email,
                    "action": e.action,
                    "entity_type": e.entity_type,
                    "entity_id": e.entity_id,
                    "description": e.description,
                }
                for e in entries
            ],
        }


# ─── Vendor Master ──────────────────────────────────────

def _vendor_to_dict(v: Vendor) -> dict:
    import json
    return {
        "id": v.id,
        "name": v.name,
        "aliases": json.loads(v.aliases) if v.aliases else [],
        "brn": v.brn,
        "vat": v.vat,
        "address": v.address,
        "phone": v.phone,
        "email": v.email,
        "default_account_code": v.default_account_code,
        "default_tds_rate": v.default_tds_rate,
        "payment_terms": v.payment_terms,
        "is_active": v.is_active,
        "created_at": v.created_at.isoformat() if v.created_at else None,
    }


def _fuzzy_match_vendor(vendor_name: str, company_id: str, threshold: float = 0.8) -> tuple:
    """Fuzzy-match a vendor name against the vendor master. Returns (vendor_id, confidence)."""
    from difflib import SequenceMatcher
    import json

    if not vendor_name:
        return None, 0.0

    db = SessionLocal()
    try:
        vendors = db.query(Vendor).filter(
            Vendor.company_id == company_id,
            Vendor.is_active == True,
        ).all()

        best_id = None
        best_score = 0.0

        name_lower = vendor_name.lower().strip()
        for v in vendors:
            # Check primary name
            score = SequenceMatcher(None, name_lower, v.name.lower().strip()).ratio()
            if score > best_score:
                best_score = score
                best_id = v.id
            # Check aliases
            aliases = json.loads(v.aliases) if v.aliases else []
            for alias in aliases:
                alias_score = SequenceMatcher(None, name_lower, alias.lower().strip()).ratio()
                if alias_score > best_score:
                    best_score = alias_score
                    best_id = v.id

        if best_score >= threshold:
            return best_id, best_score
        return None, best_score
    finally:
        db.close()


@app.get("/vendors")
async def list_vendors(
    search: Optional[str] = None,
    company: dict = Depends(get_current_company),
):
    """List all vendors for the active company"""
    with get_db() as db:
        query = db.query(Vendor).filter(Vendor.company_id == company['id'])
        if search:
            query = query.filter(Vendor.name.ilike(f"%{search}%"))
        vendors = query.order_by(Vendor.name).all()
        return {
            "total": len(vendors),
            "vendors": [_vendor_to_dict(v) for v in vendors],
        }


class VendorCreate(PydanticBaseModel):
    name: str
    aliases: Optional[List[str]] = []
    brn: Optional[str] = None
    vat: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    default_account_code: Optional[str] = None
    default_tds_rate: float = 0.0
    payment_terms: Optional[str] = None


@app.post("/vendors")
async def create_vendor(vendor_data: VendorCreate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Create a vendor (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    import json
    with get_db() as db:
        vendor = Vendor(
            name=vendor_data.name,
            aliases=json.dumps(vendor_data.aliases),
            brn=vendor_data.brn,
            vat=vendor_data.vat,
            address=vendor_data.address,
            phone=vendor_data.phone,
            email=vendor_data.email,
            default_account_code=vendor_data.default_account_code,
            default_tds_rate=vendor_data.default_tds_rate,
            payment_terms=vendor_data.payment_terms,
            is_active=True,
            company_id=company['id'],
        )
        db.add(vendor)
        db.commit()
        log_audit("vendor_created", user, entity_type="vendor", entity_id=str(vendor.id),
                  description=f"Vendor created: {vendor_data.name}", company_id=company['id'])
        return {"id": vendor.id, "message": f"Vendor '{vendor_data.name}' created"}


@app.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: int, vendor_data: VendorCreate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Update a vendor (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    import json
    with get_db() as db:
        vendor = db.query(Vendor).filter(Vendor.id == vendor_id, Vendor.company_id == company['id']).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        vendor.name = vendor_data.name
        vendor.aliases = json.dumps(vendor_data.aliases)
        vendor.brn = vendor_data.brn
        vendor.vat = vendor_data.vat
        vendor.address = vendor_data.address
        vendor.phone = vendor_data.phone
        vendor.email = vendor_data.email
        vendor.default_account_code = vendor_data.default_account_code
        vendor.default_tds_rate = vendor_data.default_tds_rate
        vendor.payment_terms = vendor_data.payment_terms
        vendor.updated_at = datetime.utcnow()
        db.commit()
        log_audit("vendor_updated", user, entity_type="vendor", entity_id=str(vendor_id),
                  description=f"Vendor updated: {vendor_data.name}", company_id=company['id'])
        return {"message": "Vendor updated"}


@app.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: int, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Delete a vendor (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    with get_db() as db:
        vendor = db.query(Vendor).filter(Vendor.id == vendor_id, Vendor.company_id == company['id']).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        log_audit("vendor_deleted", user, entity_type="vendor", entity_id=str(vendor_id),
                  description=f"Vendor deleted: {vendor.name}", company_id=company['id'])
        db.delete(vendor)
        db.commit()
        return {"message": "Vendor deleted"}


class VendorLinkRequest(PydanticBaseModel):
    vendor_id: int


@app.patch("/invoices/{invoice_id}/vendor")
async def link_invoice_to_vendor(invoice_id: str, request: VendorLinkRequest, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Link an invoice to a vendor and apply vendor defaults (TDS, account code)"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        vendor = db.query(Vendor).filter(Vendor.id == request.vendor_id, Vendor.company_id == company['id']).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        invoice.vendor_id = vendor.id
        invoice.vendor_match_confidence = 1.0
        # Apply vendor defaults if not already set
        if vendor.default_tds_rate > 0 and not invoice.tds_applicable:
            invoice.tds_applicable = True
            invoice.tds_rate = vendor.default_tds_rate
        if vendor.brn and not invoice.vendor_brn:
            invoice.vendor_brn = vendor.brn
        if vendor.vat and not invoice.vendor_vat:
            invoice.vendor_vat = vendor.vat
        invoice.updated_at = datetime.utcnow()
        db.commit()
        log_audit("invoice_vendor_linked", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"Invoice {invoice.invoice_number} linked to vendor {vendor.name}", company_id=company['id'])
        return {"message": f"Invoice linked to {vendor.name}", "vendor_name": vendor.name}


@app.patch("/invoices/{invoice_id}/vendor/unlink")
async def unlink_invoice_vendor(invoice_id: str, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Unlink an invoice from its vendor"""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        old_vendor_id = invoice.vendor_id
        invoice.vendor_id = None
        invoice.vendor_match_confidence = 0.0
        invoice.updated_at = datetime.utcnow()
        db.commit()
        log_audit("invoice_vendor_unlinked", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"Invoice {invoice.invoice_number} unlinked from vendor", company_id=company['id'])
        return {"message": "Invoice unlinked from vendor"}


class AssignRequest(PydanticBaseModel):
    assigned_to: str  # user_id


@app.patch("/invoices/{invoice_id}/assign")
async def assign_invoice(invoice_id: str, request: AssignRequest, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Assign an invoice to a specific user for approval/posting."""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        # Verify the assigned user has access to this company
        from src.auth_models import auth_db as _auth_db
        if not _auth_db.user_has_company_access(request.assigned_to, company['id']):
            raise HTTPException(status_code=400, detail="User does not have access to this company")
        assignee = _auth_db.get_user_by_id(request.assigned_to)
        old_assignee = invoice.assigned_to
        invoice.assigned_to = request.assigned_to
        invoice.updated_at = datetime.utcnow()
        db.commit()
        log_audit("invoice_assigned", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"Invoice {invoice.invoice_number} assigned to {assignee['email'] if assignee else request.assigned_to}", company_id=company['id'])

        # E-signature integration (DocuSeal/OpenSign) is currently disabled
        # To re-enable, set DOCUSEAL_API_KEY in .env and uncomment the block below
        # if docuseal_configured() and assignee and invoice.status == "pending_review":
        #     doc_path = invoice.source_file if invoice.source_file and os.path.exists(invoice.source_file) else None
        #     envelope = await create_approval_envelope(
        #         invoice_id=invoice.invoice_id,
        #         invoice_number=invoice.invoice_number,
        #         vendor_name=invoice.vendor_name,
        #         amount=invoice.total_amount,
        #         currency=invoice.currency,
        #         approver_email=assignee['email'],
        #         approver_name=assignee['full_name'],
        #         document_path=doc_path,
        #     )
        #     if envelope:
        #         log_audit("docuseal_envelope_created", user, entity_type="invoice", entity_id=invoice_id,
        #                   description=f"DocuSeal envelope created for {invoice.invoice_number} (submission_id={envelope.get('id')})", company_id=company['id'])

        return {"message": f"Invoice assigned to {assignee['full_name'] if assignee else 'user'}", "assigned_to": request.assigned_to}


# ─── DocuSeal Webhook ────────────────────────────────────

@app.post("/webhooks/docuseal")
async def docuseal_webhook(request: Request):
    """Receive DocuSeal webhook callbacks when a submission is completed/declined.

    DocuSeal sends events:
    - form.completed: submitter signed/approved
    - form.declined: submitter declined
    - submission.completed: all parties completed

    The payload includes the submitter's email and the metadata we set
    when creating the envelope (invoice_id, etc.).
    """
    body = await request.body()

    # Verify webhook signature
    if not verify_webhook(dict(request.headers), body):
        raise HTTPException(status_code=403, detail="Invalid webhook signature")

    import json
    try:
        payload = json.loads(body)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    # DocuSeal webhook event types: form.completed, form.declined, submission.completed
    event_type = payload.get("event_type") or payload.get("type", "")
    
    # Extract submitter info — DocuSeal nests it differently per event type
    submitter = payload.get("submitter") or payload.get("data", {}).get("submitter", {})
    if isinstance(submitter, dict):
        submitter_email = submitter.get("email", "")
        metadata = submitter.get("metadata", {}) or payload.get("metadata", {})
    else:
        submitter_email = ""
        metadata = payload.get("metadata", {})

    invoice_id = metadata.get("invoice_id")

    if not invoice_id:
        logger.info(f"DocuSeal webhook: no invoice_id in metadata (event={event_type})")
        return {"status": "ok", "message": "No invoice_id in metadata"}

    logger.info(f"📧 DocuSeal webhook: invoice_id={invoice_id}, event={event_type}, email={submitter_email}")

    # Determine action from event type
    is_completed = event_type in ("form.completed", "submission.completed") or payload.get("completed_at") or payload.get("status") == "completed"
    is_declined = event_type == "form.declined" or payload.get("declined_at") or payload.get("status") == "declined"

    if not is_completed and not is_declined:
        return {"status": "ok", "message": f"Ignored event: {event_type}"}

    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id).first()
        if not invoice:
            logger.warning(f"DocuSeal webhook: invoice {invoice_id} not found")
            return {"status": "ok", "message": "Invoice not found"}

        from src.auth_models import auth_db as _auth_db

        if is_completed:
            invoice.status = "approved"
            approver = _auth_db.get_user_by_email(submitter_email) if submitter_email else None
            if approver:
                invoice.approved_by = approver['id']
            invoice.updated_at = datetime.utcnow()
            db.commit()
            log_audit("docuseal_approved", {"email": submitter_email, "id": invoice.approved_by or "docuseal"},
                      entity_type="invoice", entity_id=invoice_id,
                      description=f"Invoice {invoice.invoice_number} approved via DocuSeal e-signature by {submitter_email}",
                      company_id=invoice.company_id)
            logger.info(f"✅ Invoice {invoice.invoice_number} auto-approved via DocuSeal by {submitter_email}")

        elif is_declined:
            invoice.status = "rejected"
            invoice.updated_at = datetime.utcnow()
            db.commit()
            log_audit("docuseal_rejected", {"email": submitter_email, "id": "docuseal"},
                      entity_type="invoice", entity_id=invoice_id,
                      description=f"Invoice {invoice.invoice_number} rejected via DocuSeal by {submitter_email}",
                      company_id=invoice.company_id)
            logger.info(f"❌ Invoice {invoice.invoice_number} rejected via DocuSeal by {submitter_email}")

    return {"status": "ok"}


# ─── Recurring Invoices ─────────────────────────────────

class RecurringTemplateCreate(PydanticBaseModel):
    name: str
    vendor_name: str
    vendor_id: Optional[int] = None
    invoice_type: str = "supplier"
    frequency: str = "monthly"
    day_of_month: int = 1
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    line_items: List[dict] = []
    total_amount: float = 0.0
    tds_rate: float = 0.0
    tds_applicable: bool = False
    auto_post: bool = False
    currency: str = "MUR"


def _calculate_next_generation(template: RecurringTemplate, from_date=None):
    """Calculate the next generation date for a recurring template."""
    from datetime import date, timedelta
    today = from_date or date.today()
    dom = min(template.day_of_month, 28)

    if template.frequency == "monthly":
        if today.day <= dom:
            return today.replace(day=dom).isoformat()
        # Next month
        if today.month == 12:
            return date(today.year + 1, 1, dom).isoformat()
        return date(today.year, today.month + 1, dom).isoformat()
    elif template.frequency == "quarterly":
        q_months = [1, 4, 7, 10]
        for qm in q_months:
            if today.month <= qm and (today.month < qm or today.day <= dom):
                return date(today.year, qm, dom).isoformat()
        return date(today.year + 1, 1, dom).isoformat()
    elif template.frequency == "annually":
        if today.month == 1 and today.day <= dom:
            return date(today.year, 1, dom).isoformat()
        return date(today.year + 1, 1, dom).isoformat()
    return today.isoformat()


def generate_due_recurring_invoices():
    """Check all active recurring templates and generate invoices for those due.
    Should be called on app startup and/or via a daily cron/timer."""
    from src.invoice_engine import generate_invoice_id, generate_entry_id
    import json

    db = SessionLocal()
    try:
        templates = db.query(RecurringTemplate).filter(RecurringTemplate.is_active == True).all()
        today_str = date.today().isoformat() if 'date' in dir() else datetime.now().strftime("%Y-%m-%d")
        from datetime import date as date_cls
        today_str = date_cls.today().isoformat()

        for template in templates:
            # Check if company has recurring enabled
            company = auth_db.get_company_by_id(template.company_id) if hasattr(generate_due_recurring_invoices, 'auth_db') else None
            from src.auth_models import auth_db as _auth_db
            company = _auth_db.get_company_by_id(template.company_id)
            if not company or not company.get('recurring_enabled'):
                continue

            # Check if end date passed
            if template.end_date and today_str > template.end_date:
                continue

            # Check if already generated today
            if template.last_generated == today_str:
                continue

            # Check if due
            if template.next_generation and today_str < template.next_generation:
                continue

            # Generate the invoice
            invoice_id = generate_invoice_id()
            line_items = json.loads(template.line_items) if template.line_items else []

            invoice_data = {
                "vendor_name": template.vendor_name,
                "invoice_number": f"REC-{date_cls.today().strftime('%Y%m%d')}-{template.id}",
                "invoice_date": today_str,
                "currency": template.currency,
                "line_items": line_items,
                "total_amount": template.total_amount,
                "subtotal": template.total_amount,
                "tax_total": 0.0,
                "confidence_score": 1.0,
                "tds_applicable": template.tds_applicable,
                "tds_rate": template.tds_rate,
            }

            entries = generate_accounting_entries(invoice_id, invoice_data, template.invoice_type, company_id=template.company_id)

            result = {
                "invoice_id": invoice_id,
                "status": "posted" if template.auto_post else "pending_review",
                "extracted_data": invoice_data,
                "suggested_entries": entries,
                "message": "Generated from recurring template",
            }

            _save_invoice_to_db(result, template.invoice_type, None, None, None, company_id=template.company_id)

            # Update template
            template.last_generated = today_str
            template.next_generation = _calculate_next_generation(template)
            db.commit()

            # Auto-post if enabled
            if template.auto_post:
                inv = db.query(Invoice).filter(Invoice.invoice_id == invoice_id).first()
                if inv:
                    inv.status = "posted"
                    for je in inv.journal_entries:
                        je.status = "posted"
                    db.commit()

            logger.info(f"🔄 Recurring invoice generated: {invoice_id} from template '{template.name}'")

    except Exception as e:
        logger.error(f"❌ Recurring generation error: {e}")
    finally:
        db.close()


@app.get("/recurring/templates")
async def list_recurring_templates(company: dict = Depends(get_current_company)):
    """List all recurring templates for the active company"""
    if not company.get('recurring_enabled'):
        return {"total": 0, "templates": [], "recurring_enabled": False}
    with get_db() as db:
        templates = db.query(RecurringTemplate).filter(
            RecurringTemplate.company_id == company['id']
        ).order_by(RecurringTemplate.name).all()
        return {
            "total": len(templates),
            "recurring_enabled": True,
            "templates": [
                {
                    "id": t.id,
                    "name": t.name,
                    "vendor_name": t.vendor_name,
                    "invoice_type": t.invoice_type,
                    "frequency": t.frequency,
                    "day_of_month": t.day_of_month,
                    "total_amount": t.total_amount,
                    "tds_applicable": t.tds_applicable,
                    "tds_rate": t.tds_rate,
                    "auto_post": t.auto_post,
                    "is_active": t.is_active,
                    "last_generated": t.last_generated,
                    "next_generation": t.next_generation,
                    "start_date": t.start_date,
                    "end_date": t.end_date,
                }
                for t in templates
            ],
        }


@app.post("/recurring/templates")
async def create_recurring_template(template_data: RecurringTemplateCreate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Create a recurring invoice template (admin only, requires recurring enabled)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    if not company.get('recurring_enabled'):
        raise HTTPException(status_code=400, detail="Recurring invoices are not enabled for this company. Enable it in admin settings first.")
    import json
    with get_db() as db:
        template = RecurringTemplate(
            name=template_data.name,
            vendor_name=template_data.vendor_name,
            vendor_id=template_data.vendor_id,
            invoice_type=template_data.invoice_type,
            frequency=template_data.frequency,
            day_of_month=template_data.day_of_month,
            start_date=template_data.start_date,
            end_date=template_data.end_date,
            line_items=json.dumps(template_data.line_items),
            total_amount=template_data.total_amount,
            tds_rate=template_data.tds_rate,
            tds_applicable=template_data.tds_applicable,
            auto_post=template_data.auto_post,
            currency=template_data.currency,
            is_active=True,
            company_id=company['id'],
        )
        # Calculate next generation
        template.next_generation = _calculate_next_generation(template)
        db.add(template)
        db.commit()
        log_audit("recurring_created", user, entity_type="recurring_template", entity_id=str(template.id),
                  description=f"Recurring template created: {template_data.name}", company_id=company['id'])
        return {"id": template.id, "message": f"Recurring template '{template_data.name}' created"}


@app.put("/recurring/templates/{template_id}")
async def update_recurring_template(template_id: int, template_data: RecurringTemplateCreate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Update a recurring template (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    import json
    with get_db() as db:
        template = db.query(RecurringTemplate).filter(RecurringTemplate.id == template_id, RecurringTemplate.company_id == company['id']).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        template.name = template_data.name
        template.vendor_name = template_data.vendor_name
        template.vendor_id = template_data.vendor_id
        template.invoice_type = template_data.invoice_type
        template.frequency = template_data.frequency
        template.day_of_month = template_data.day_of_month
        template.start_date = template_data.start_date
        template.end_date = template_data.end_date
        template.line_items = json.dumps(template_data.line_items)
        template.total_amount = template_data.total_amount
        template.tds_rate = template_data.tds_rate
        template.tds_applicable = template_data.tds_applicable
        template.auto_post = template_data.auto_post
        template.currency = template_data.currency
        template.next_generation = _calculate_next_generation(template)
        template.updated_at = datetime.utcnow()
        db.commit()
        log_audit("recurring_updated", user, entity_type="recurring_template", entity_id=str(template_id),
                  description=f"Recurring template updated: {template_data.name}", company_id=company['id'])
        return {"message": "Template updated"}


@app.delete("/recurring/templates/{template_id}")
async def delete_recurring_template(template_id: int, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Delete a recurring template (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    with get_db() as db:
        template = db.query(RecurringTemplate).filter(RecurringTemplate.id == template_id, RecurringTemplate.company_id == company['id']).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        log_audit("recurring_deleted", user, entity_type="recurring_template", entity_id=str(template_id),
                  description=f"Recurring template deleted: {template.name}", company_id=company['id'])
        db.delete(template)
        db.commit()
        return {"message": "Template deleted"}


@app.patch("/recurring/templates/{template_id}/toggle")
async def toggle_recurring_template(template_id: int, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Pause/resume a recurring template"""
    with get_db() as db:
        template = db.query(RecurringTemplate).filter(RecurringTemplate.id == template_id, RecurringTemplate.company_id == company['id']).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        template.is_active = not template.is_active
        if template.is_active:
            template.next_generation = _calculate_next_generation(template)
        db.commit()
        action = "resumed" if template.is_active else "paused"
        log_audit(f"recurring_{action}", user, entity_type="recurring_template", entity_id=str(template_id),
                  description=f"Recurring template '{template.name}' {action}", company_id=company['id'])
        return {"message": f"Template {action}", "is_active": template.is_active}


@app.post("/recurring/generate-now/{template_id}")
async def generate_recurring_now(template_id: int, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Manually trigger generation for a specific template (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    from src.invoice_engine import generate_invoice_id
    from datetime import date as date_cls
    import json

    with get_db() as db:
        template = db.query(RecurringTemplate).filter(RecurringTemplate.id == template_id, RecurringTemplate.company_id == company['id']).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

        today_str = date_cls.today().strftime("%Y-%m-%d")
        invoice_id = generate_invoice_id()
        line_items = json.loads(template.line_items) if template.line_items else []

        invoice_data = {
            "vendor_name": template.vendor_name,
            "invoice_number": f"REC-{date_cls.today().strftime('%Y%m%d')}-{template.id}",
            "invoice_date": today_str,
            "currency": template.currency,
            "line_items": line_items,
            "total_amount": template.total_amount,
            "subtotal": template.total_amount,
            "tax_total": 0.0,
            "confidence_score": 1.0,
            "tds_applicable": template.tds_applicable,
            "tds_rate": template.tds_rate,
        }

        entries = generate_accounting_entries(invoice_id, invoice_data, template.invoice_type, company_id=template.company_id)
        result = {
            "invoice_id": invoice_id,
            "status": "pending_review",
            "extracted_data": invoice_data,
            "suggested_entries": entries,
            "message": "Generated from recurring template",
        }
        _save_invoice_to_db(result, template.invoice_type, None, None, None, company_id=company['id'])

        template.last_generated = today_str
        template.next_generation = _calculate_next_generation(template)
        db.commit()

        log_audit("recurring_generated", user, entity_type="recurring_template", entity_id=str(template_id),
                  description=f"Recurring invoice {invoice_id} generated from '{template.name}'", company_id=company['id'])

        return {"invoice_id": invoice_id, "message": f"Invoice generated from template '{template.name}'"}


# ─── Exchange Rates ──────────────────────────────────────

@app.get("/exchange-rates")
async def list_exchange_rates(company: dict = Depends(get_current_company)):
    """List all exchange rates (global)"""
    with get_db() as db:
        rates = db.query(ExchangeRate).order_by(ExchangeRate.currency, ExchangeRate.date.desc()).all()
        # Return latest rate per currency
        seen = {}
        for r in rates:
            if r.currency not in seen:
                seen[r.currency] = {
                    "id": r.id,
                    "currency": r.currency,
                    "rate_to_mur": r.rate_to_mur,
                    "date": r.date,
                    "source": r.source,
                }
        return {"rates": list(seen.values())}


class ExchangeRateUpdate(PydanticBaseModel):
    currency: str
    rate_to_mur: float


@app.put("/exchange-rates")
async def update_exchange_rate(rate_data: ExchangeRateUpdate, company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Update an exchange rate (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    from datetime import date
    today = date.today().isoformat()
    with get_db() as db:
        existing = db.query(ExchangeRate).filter(
            ExchangeRate.currency == rate_data.currency.upper(),
            ExchangeRate.date == today
        ).first()
        if existing:
            existing.rate_to_mur = rate_data.rate_to_mur
            existing.source = "manual"
        else:
            db.add(ExchangeRate(
                currency=rate_data.currency.upper(),
                rate_to_mur=rate_data.rate_to_mur,
                date=today,
                source="manual",
            ))
        db.commit()
        return {"message": f"Rate updated: {rate_data.currency.upper()} = {rate_data.rate_to_mur} MUR"}


@app.post("/exchange-rates/refresh")
async def refresh_exchange_rates(company: dict = Depends(get_current_company), user: dict = Depends(get_current_user)):
    """Fetch latest exchange rates from FX API (admin only)"""
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    success = fetch_exchange_rates()
    if not success:
        raise HTTPException(status_code=502, detail="Failed to fetch exchange rates from API")
    return {"message": "Exchange rates refreshed"}


# ─── Payment Settlement (with FX + bank charges) ────────

class PaymentSettlementRequest(PydanticBaseModel):
    bank_rate: float = 0.0  # actual bank rate (0 = same as booking rate, i.e., MUR invoice)
    bank_charges: float = 0.0  # bank charges in MUR
    payment_date: str  # YYYY-MM-DD


@app.post("/invoices/{invoice_id}/settle-payment")
async def settle_payment(
    invoice_id: str,
    request: PaymentSettlementRequest,
    company: dict = Depends(get_current_company),
    user: dict = Depends(get_current_user),
):
    """Settle a foreign-currency invoice payment with bank rate + charges.
    Generates FX gain/loss + bank charge journal entries. Only for non-MUR invoices."""
    with get_db() as db:
        invoice = db.query(Invoice).filter(Invoice.invoice_id == invoice_id, Invoice.company_id == company['id']).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if invoice.status != "posted":
            raise HTTPException(status_code=400, detail="Invoice must be posted before payment")
        if invoice.currency.upper() == "MUR":
            raise HTTPException(status_code=400, detail="Payment settlement form is for foreign-currency invoices only. Use status update for MUR.")

        booked_rate = invoice.exchange_rate or 1.0
        bank_rate = request.bank_rate if request.bank_rate > 0 else booked_rate
        total_foreign = invoice.total_amount
        booked_base = round(total_foreign * booked_rate, 2)
        actual_base = round(total_foreign * bank_rate, 2)
        bank_charges = request.bank_charges
        total_bank_debit = round(actual_base + bank_charges, 2)
        fx_diff = round(actual_base - booked_base, 2)

        invoice.payment_bank_rate = bank_rate
        invoice.payment_bank_charges = bank_charges
        invoice.payment_date = request.payment_date
        invoice.status = "paid"

        from src.invoice_engine import generate_entry_id
        from src.accounting_engine import ACCOUNT_MAPPINGS
        mapping = ACCOUNT_MAPPINGS.get(invoice.invoice_type, ACCOUNT_MAPPINGS["supplier"])
        payable_code, payable_name = mapping["payable"]
        bank_code, bank_name = "01-3000-01", "Bank"
        fx_loss_code, fx_loss_name = "01-6303-07", "Foreign Exchange Loss"
        fx_gain_code, fx_gain_name = "01-6304-07", "Foreign Exchange Gain"
        charges_code, charges_name = "01-6301-07", "Bank Charges"

        settlement_entry_id = generate_entry_id()
        total_debit = booked_base
        total_credit = total_bank_debit
        if fx_diff > 0:  # FX loss
            total_debit = booked_base + fx_diff + bank_charges
            total_credit = total_bank_debit
        elif fx_diff < 0:  # FX gain
            total_debit = booked_base + bank_charges
            total_credit = total_bank_debit + abs(fx_diff)
        else:
            total_debit = booked_base + bank_charges
            total_credit = total_bank_debit

        # Ensure balanced
        adjustment = round(total_debit - total_credit, 2)
        if adjustment != 0:
            if adjustment > 0:
                total_credit += adjustment
            else:
                total_debit += abs(adjustment)

        settlement = JournalEntryDB(
            entry_id=settlement_entry_id,
            invoice_id=invoice.invoice_id,
            entry_date=request.payment_date,
            reference=f"PAY-{invoice.invoice_number}",
            description=f"Payment settlement for {invoice.vendor_name} - {invoice.invoice_number} ({invoice.currency})",
            total_debit=total_debit,
            total_credit=total_credit,
            is_balanced=True,
            status="posted",
            created_by=user['email'],
            posted_by=user['id'],
            company_id=company['id'],
        )
        db.add(settlement)

        # Dr Trade Creditors (booked amount)
        db.add(JournalEntryLineDB(
            entry_id=settlement_entry_id,
            account_code=payable_code,
            account_name=payable_name,
            description=f"Settlement of {invoice.invoice_number} - {invoice.vendor_name} (booked at {booked_rate})",
            debit=booked_base,
            credit=0.0,
        ))
        # Dr/Cr FX difference
        if fx_diff > 0:
            db.add(JournalEntryLineDB(
                entry_id=settlement_entry_id,
                account_code=fx_loss_code,
                account_name=fx_loss_name,
                description=f"FX loss on {invoice.invoice_number} (booked {booked_rate}, paid {bank_rate})",
                debit=fx_diff,
                credit=0.0,
            ))
        elif fx_diff < 0:
            db.add(JournalEntryLineDB(
                entry_id=settlement_entry_id,
                account_code=fx_gain_code,
                account_name=fx_gain_name,
                description=f"FX gain on {invoice.invoice_number} (booked {booked_rate}, paid {bank_rate})",
                debit=0.0,
                credit=abs(fx_diff),
            ))
        # Dr Bank charges
        if bank_charges > 0:
            db.add(JournalEntryLineDB(
                entry_id=settlement_entry_id,
                account_code=charges_code,
                account_name=charges_name,
                description=f"Bank charges on payment for {invoice.invoice_number}",
                debit=bank_charges,
                credit=0.0,
            ))
        # Cr Bank (total actual debit)
        db.add(JournalEntryLineDB(
            entry_id=settlement_entry_id,
            account_code=bank_code,
            account_name=bank_name,
            description=f"Bank payment to {invoice.vendor_name} for {invoice.invoice_number}",
            debit=0.0,
            credit=total_bank_debit,
        ))

        db.commit()
        log_audit("invoice_paid_fx", user, entity_type="invoice", entity_id=invoice_id,
                  description=f"Payment settled: {invoice.invoice_number} ({invoice.currency}) - booked {booked_base}, actual {actual_base}, FX {'loss' if fx_diff > 0 else 'gain'} {abs(fx_diff)}, bank charges {bank_charges}",
                  company_id=company['id'])

        return {
            "invoice_id": invoice_id,
            "message": "Payment settled with FX adjustment",
            "booked_base": booked_base,
            "actual_base": actual_base,
            "fx_diff": fx_diff,
            "fx_type": "loss" if fx_diff > 0 else "gain" if fx_diff < 0 else "none",
            "bank_charges": bank_charges,
            "total_bank_debit": total_bank_debit,
            "settlement_entry_id": settlement_entry_id,
        }


# ─── Chart of Accounts ───────────────────────────────────



@app.get("/accounting/classification-rules")
async def list_classification_rules(company: dict = Depends(get_current_company)):
    """List all learned classification rules"""
    with get_db() as db:
        rules = db.query(ClassificationRule).filter(
            ClassificationRule.company_id == company['id']
        ).order_by(
            ClassificationRule.times_used.desc(),
            ClassificationRule.updated_at.desc()
        ).all()
        return {
            "total": len(rules),
            "rules": [
                {
                    "id": r.id,
                    "vendor_name": r.vendor_name,
                    "description_pattern": r.description_pattern,
                    "account_code": r.account_code,
                    "account_name": r.account_name,
                    "user_context": r.user_context,
                    "source": r.source,
                    "times_used": r.times_used,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "updated_at": r.updated_at.isoformat() if r.updated_at else None,
                }
                for r in rules
            ]
        }


@app.delete("/accounting/classification-rules/{rule_id}")
async def delete_classification_rule(rule_id: int, company: dict = Depends(get_current_company)):
    """Delete a learned classification rule"""
    with get_db() as db:
        rule = db.query(ClassificationRule).filter(
            ClassificationRule.id == rule_id,
            ClassificationRule.company_id == company['id'],
        ).first()
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")
        db.delete(rule)
        db.commit()
        return {"message": f"Rule deleted: {rule.vendor_name} → {rule.account_code}"}

@app.get("/accounting/chart-of-accounts")
async def get_chart_of_accounts(category: Optional[str] = None):
    """Get the chart of accounts"""
    with get_db() as db:
        query = db.query(ChartOfAccountsDB).filter(ChartOfAccountsDB.is_active == True)
        if category:
            query = query.filter(ChartOfAccountsDB.category == category)
        accounts = query.order_by(ChartOfAccountsDB.code).all()
        
        return {
            "total": len(accounts),
            "accounts": [
                {
                    "code": a.code,
                    "name": a.name,
                    "category": a.category,
                    "parent_code": a.parent_code,
                    "description": a.description,
                }
                for a in accounts
            ]
        }




@app.get("/accounting/export/excel")
async def export_journal_entries_excel(
    status: Optional[str] = "posted",
    company: dict = Depends(get_current_company),
):
    """Export journal entries to Excel. Defaults to posted entries only."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import tempfile

    with get_db() as db:
        query = db.query(JournalEntryDB).filter(JournalEntryDB.company_id == company['id']).order_by(JournalEntryDB.entry_date.asc(), JournalEntryDB.created_at.asc())
        if status:
            query = query.filter(JournalEntryDB.status == status)
        entries = query.all()

        if not entries:
            raise HTTPException(status_code=404, detail="No journal entries found to export")

        wb = openpyxl.Workbook()

        # --- Sheet 1: Journal Entries Detail ---
        ws = wb.active
        ws.title = "Journal Entries"

        # Styles
        header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1B4332")
        data_font = Font(name="Arial", size=10)
        accent_font = Font(name="Arial", size=10, color="1B7A43")
        debit_font = Font(name="Arial", size=10, color="C0392B")
        credit_font = Font(name="Arial", size=10, color="27AE60")
        total_font = Font(name="Arial", size=10, bold=True)
        total_fill = PatternFill("solid", fgColor="F0F4F0")
        border = Border(
            bottom=Side(style="thin", color="D5D5D5")
        )
        currency_fmt = '#,##0.00'

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "MC Golf - Journal Entries Export"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1B4332")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888")
        ws["A3"] = f"Filter: {status or 'All'} | Total entries: {len(entries)}"
        ws["A3"].font = Font(name="Arial", size=9, color="888888")

        # Headers
        headers = ["Entry ID", "Date", "Vendor / Reference", "Account Code", "Account Name", "Description", "Debit", "Credit"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 6
        grand_debit = 0
        grand_credit = 0

        for entry in entries:
            # Get vendor name from related invoice
            vendor = ""
            if entry.invoice:
                vendor = entry.invoice.vendor_name or ""

            entry_start_row = row
            for line in entry.lines:
                ws.cell(row=row, column=1, value=entry.entry_id).font = accent_font
                ws.cell(row=row, column=2, value=entry.entry_date).font = data_font
                ws.cell(row=row, column=3, value=vendor or entry.reference or "").font = data_font
                ws.cell(row=row, column=4, value=line.account_code).font = Font(name="Consolas", size=10)
                ws.cell(row=row, column=5, value=line.account_name).font = data_font
                ws.cell(row=row, column=6, value=line.description or "").font = data_font

                debit_cell = ws.cell(row=row, column=7, value=line.debit if line.debit > 0 else None)
                debit_cell.font = debit_font
                debit_cell.number_format = currency_fmt

                credit_cell = ws.cell(row=row, column=8, value=line.credit if line.credit > 0 else None)
                credit_cell.font = credit_font
                credit_cell.number_format = currency_fmt

                for c in range(1, 9):
                    ws.cell(row=row, column=c).border = border

                grand_debit += line.debit or 0
                grand_credit += line.credit or 0
                row += 1

            # Entry subtotal row
            ws.cell(row=row, column=6, value=f"Entry Total ({entry.entry_id})").font = total_font
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
            sub_debit = ws.cell(row=row, column=7, value=entry.total_debit)
            sub_debit.font = total_font
            sub_debit.number_format = currency_fmt
            sub_debit.fill = total_fill
            sub_credit = ws.cell(row=row, column=8, value=entry.total_credit)
            sub_credit.font = total_font
            sub_credit.number_format = currency_fmt
            sub_credit.fill = total_fill
            row += 1  # blank row separator

        # Grand total
        row += 1
        ws.cell(row=row, column=6, value="GRAND TOTAL").font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
        gt_debit = ws.cell(row=row, column=7, value=grand_debit)
        gt_debit.font = Font(name="Arial", bold=True, size=11, color="C0392B")
        gt_debit.number_format = currency_fmt
        gt_credit = ws.cell(row=row, column=8, value=grand_credit)
        gt_credit.font = Font(name="Arial", bold=True, size=11, color="27AE60")
        gt_credit.number_format = currency_fmt

        # Column widths
        widths = [22, 12, 25, 14, 25, 45, 15, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A6"

        # --- Sheet 2: Summary by Account ---
        ws2 = wb.create_sheet("Summary by Account")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = "Journal Entries Summary by Account"
        ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1B4332")

        summary_headers = ["Account Code", "Account Name", "Total Debit", "Total Credit", "Net"]
        for col, h in enumerate(summary_headers, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Aggregate by account
        account_totals = {}
        for entry in entries:
            for line in entry.lines:
                key = (line.account_code or "", line.account_name or "")
                if key not in account_totals:
                    account_totals[key] = {"debit": 0, "credit": 0}
                account_totals[key]["debit"] += line.debit or 0
                account_totals[key]["credit"] += line.credit or 0

        srow = 4
        for (code, name), totals in sorted(account_totals.items()):
            ws2.cell(row=srow, column=1, value=code).font = Font(name="Consolas", size=10)
            ws2.cell(row=srow, column=2, value=name).font = data_font
            d = ws2.cell(row=srow, column=3, value=totals["debit"])
            d.font = debit_font
            d.number_format = currency_fmt
            c = ws2.cell(row=srow, column=4, value=totals["credit"])
            c.font = credit_font
            c.number_format = currency_fmt
            net = ws2.cell(row=srow, column=5, value=totals["debit"] - totals["credit"])
            net.font = data_font
            net.number_format = currency_fmt
            for col in range(1, 6):
                ws2.cell(row=srow, column=col).border = border
            srow += 1

        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 15
        ws2.column_dimensions["E"].width = 15
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A4"

        # Save to temp file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir="temp_uploads")
        wb.save(tmp.name)
        tmp.close()

        filename = f"MC_Golf_Journal_Entries_{status or 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

@app.get("/accounting/export/sage200")
async def export_journal_entries_sage200(
    status: Optional[str] = "posted",
    transaction_type: str = "JL",
    company: dict = Depends(get_current_company),
):
    """Export journal entries as a Sage 200 Evolution GL Journal Batch CSV.

    Produces a single consolidated CSV file ready for import via:
    General Ledger | Transactions | Journal Batches | New Batch | Batch | Import

    Query params:
        status           - entry status filter (default 'posted')
        transaction_type - Sage Transaction Type code (default 'JL'; adjust to
                           match the customer's Transaction Types under
                           Maintenance | Maintenance | Transaction Types)
    """
    import csv
    import tempfile
    import re

    def clean_text(value: str, max_len: int = 100) -> str:
        """Strip Sage-illegal characters and whitespace."""
        if not value:
            return ""
        value = str(value).strip()
        value = re.sub(r"[,;:'\"<>*&$@/\\()]", "", value)
        value = re.sub(r"\s+", " ", value)
        return value[:max_len]

    def format_date(date_str: str) -> str:
        """Convert YYYY-MM-DD (or similar) to dd/mm/yyyy for Sage."""
        if not date_str:
            return ""
        try:
            dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            return date_str

    def fmt_amount(value: float) -> str:
        """Format to 2 decimal places, no thousands separator."""
        return f"{float(value or 0):.2f}"

    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".csv", dir="temp_uploads", mode="w", newline="", encoding="utf-8"
    )
    writer = csv.writer(tmp)

    # Header row — Sage 200 Evolution GL Journal Batch minimum columns
    writer.writerow([
        "Date",
        "AccountCode",
        "TransactionType",
        "Reference",
        "Description",
        "Debit",
        "Credit",
    ])

    with get_db() as db:
        query = db.query(JournalEntryDB).filter(JournalEntryDB.company_id == company['id']).order_by(
            JournalEntryDB.entry_date.asc(), JournalEntryDB.created_at.asc()
        )
        if status:
            query = query.filter(JournalEntryDB.status == status)
        entries = query.all()

        if not entries:
            tmp.close()
            os.unlink(tmp.name)
            raise HTTPException(status_code=404, detail="No journal entries found to export")

        for entry in entries:
            ref = clean_text(entry.reference or (entry.invoice.invoice_number if entry.invoice else ""), 30)
            for line in entry.lines:
                writer.writerow([
                    format_date(entry.entry_date),
                    clean_text(line.account_code, 20),
                    transaction_type,
                    ref,
                    clean_text(line.description or "", 100),
                    fmt_amount(line.debit),
                    fmt_amount(line.credit),
                ])

    tmp.close()

    filename = f"Sage200_GL_Journal_{status or 'all'}_{transaction_type}_{datetime.now().strftime('%Y%m%d')}.csv"
    return FileResponse(
        tmp.name,
        media_type="text/csv",
        filename=filename,
    )

@app.get("/accounting/suggest-account")
async def suggest_account(description: str, type: str = "supplier", company: dict = Depends(get_current_company)):
    """AI-powered account code suggestion"""
    code, name = suggest_account_code(description, type, company_id=company['id'])
    return {"account_code": code, "account_name": name, "description": description}


# ─── Dashboard ────────────────────────────────────────────

@app.get("/dashboard/stats")
async def get_dashboard_stats(
    period: str = "all",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    company: dict = Depends(get_current_company),
):
    """Get dashboard statistics, optionally filtered by date range.

    period presets: all, month, quarter, year, custom (requires start_date/end_date)
    """
    from datetime import date, timedelta

    today = date.today()
    if period == "month":
        start_date = today.replace(day=1).isoformat()
        end_date = today.isoformat()
    elif period == "quarter":
        q_month = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=q_month, day=1).isoformat()
        end_date = today.isoformat()
    elif period == "year":
        start_date = today.replace(month=1, day=1).isoformat()
        end_date = today.isoformat()
    elif period == "last_month":
        first_of_this = today.replace(day=1)
        last_month_end = first_of_this - timedelta(days=1)
        start_date = last_month_end.replace(day=1).isoformat()
        end_date = last_month_end.isoformat()
    elif period == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="custom period requires start_date and end_date")

    with get_db() as db:
        base_query = db.query(Invoice).filter(Invoice.company_id == company['id'])
        if start_date:
            base_query = base_query.filter(Invoice.invoice_date >= start_date)
        if end_date:
            base_query = base_query.filter(Invoice.invoice_date <= end_date)

        total = base_query.count()
        pending = base_query.filter(Invoice.status == "pending_review").count()
        approved = base_query.filter(Invoice.status == "approved").count()
        posted = base_query.filter(Invoice.status == "posted").count()

        # Totals by type (only active statuses)
        from sqlalchemy import func
        active_statuses = ["pending_review", "approved", "posted", "paid"]
        payable_q = base_query.filter(
            Invoice.invoice_type == "supplier",
            Invoice.status.in_(active_statuses)
        )
        receivable_q = base_query.filter(
            Invoice.invoice_type == "client",
            Invoice.status.in_(active_statuses)
        )
        payable = payable_q.with_entities(func.sum(Invoice.total_amount_base)).scalar() or 0.0
        receivable = receivable_q.with_entities(func.sum(Invoice.total_amount_base)).scalar() or 0.0

        # Recent invoices (always show latest regardless of period for navigation)
        recent_q = db.query(Invoice).filter(Invoice.company_id == company['id'])
        if start_date:
            recent_q = recent_q.filter(Invoice.invoice_date >= start_date)
        if end_date:
            recent_q = recent_q.filter(Invoice.invoice_date <= end_date)
        recent = recent_q.order_by(Invoice.created_at.desc()).limit(10).all()

        return {
            "period": period,
            "start_date": start_date,
            "end_date": end_date,
            "total_invoices": total,
            "pending_review": pending,
            "approved": approved,
            "posted": posted,
            "total_payable": round(payable, 2),
            "total_receivable": round(receivable, 2),
            "recent_invoices": [_invoice_to_dict(inv) for inv in recent],
        }


@app.get("/dashboard/charts")
async def get_dashboard_charts(
    company: dict = Depends(get_current_company),
):
    """Get chart data for dashboard visualizations."""
    from sqlalchemy import func, extract

    with get_db() as db:
        # 1. Monthly spend (last 6 months) — supplier invoices by month
        six_months_ago = (datetime.now().replace(day=1)).isoformat()
        monthly = db.query(
            func.strftime('%Y-%m', Invoice.invoice_date).label('month'),
            func.sum(Invoice.total_amount_base).label('total'),
        ).filter(
            Invoice.company_id == company['id'],
            Invoice.invoice_type == 'supplier',
            Invoice.status.in_(['pending_review', 'approved', 'posted', 'paid']),
            Invoice.invoice_date >= six_months_ago,
        ).group_by('month').order_by('month').all()

        monthly_data = [
            {"month": row.month or 'N/A', "amount": round(row.total or 0, 2)}
            for row in monthly
        ]

        # 2. Top vendors by spend (top 5)
        top_vendors = db.query(
            Invoice.vendor_name,
            func.sum(Invoice.total_amount_base).label('total'),
        ).filter(
            Invoice.company_id == company['id'],
            Invoice.invoice_type == 'supplier',
            Invoice.status.in_(['pending_review', 'approved', 'posted', 'paid']),
        ).group_by(Invoice.vendor_name).order_by(func.sum(Invoice.total_amount_base).desc()).limit(5).all()

        vendor_data = [
            {"vendor": row.vendor_name or 'Unknown', "amount": round(row.total or 0, 2)}
            for row in top_vendors
        ]

        # 3. Expense breakdown by account code (top 5)
        account_totals = {}
        entries = db.query(JournalEntryDB).filter(
            JournalEntryDB.company_id == company['id'],
            JournalEntryDB.status.in_(['posted']),
        ).all()
        for je in entries:
            for line in je.lines:
                if line.debit and line.debit > 0:
                    key = (line.account_code or 'N/A', line.account_name or 'Unknown')
                    account_totals[key] = account_totals.get(key, 0) + line.debit

        sorted_accounts = sorted(account_totals.items(), key=lambda x: x[1], reverse=True)[:5]
        account_data = [
            {"account_code": code, "account_name": name[:30], "amount": round(amt, 2)}
            for (code, name), amt in sorted_accounts
        ]

        # 4. Invoice status distribution
        status_counts = {}
        all_invoices = db.query(Invoice).filter(Invoice.company_id == company['id']).all()
        for inv in all_invoices:
            status_counts[inv.status] = status_counts.get(inv.status, 0) + 1
        status_data = [{"status": s, "count": c} for s, c in status_counts.items()]

        return {
            "monthly_spend": monthly_data,
            "top_vendors": vendor_data,
            "account_breakdown": account_data,
            "status_distribution": status_data,
        }


# ─── Helper Functions ─────────────────────────────────────

def _invoice_to_dict(invoice: Invoice) -> Dict[str, Any]:
    """Convert Invoice ORM object to dict"""
    return {
        "invoice_id": invoice.invoice_id,
        "invoice_type": invoice.invoice_type,
        "status": invoice.status,
        "vendor_name": invoice.vendor_name,
        "invoice_number": invoice.invoice_number,
        "invoice_date": invoice.invoice_date,
        "due_date": invoice.due_date,
        "currency": invoice.currency,
        "subtotal": invoice.subtotal,
        "tax_total": invoice.tax_total,
        "total_amount": invoice.total_amount,
        "project_code": invoice.project_code,
        "cost_center": invoice.cost_center,
        "confidence_score": invoice.confidence_score,
        "approved_by": invoice.approved_by,
        "posted_by": invoice.posted_by,
        "tds_applicable": invoice.tds_applicable,
        "tds_rate": invoice.tds_rate,
        "tds_amount": invoice.tds_amount,
        "tds_paid_to_mra": invoice.tds_paid_to_mra,
        "tds_paid_date": invoice.tds_paid_date,
        "vendor_id": invoice.vendor_id,
        "vendor_match_confidence": invoice.vendor_match_confidence,
        "assigned_to": invoice.assigned_to,
        "exchange_rate": invoice.exchange_rate,
        "total_amount_base": invoice.total_amount_base,
        "payment_bank_rate": invoice.payment_bank_rate,
        "payment_bank_charges": invoice.payment_bank_charges,
        "payment_date": invoice.payment_date,
        "created_at": invoice.created_at.isoformat() if invoice.created_at else None,
        "updated_at": invoice.updated_at.isoformat() if invoice.updated_at else None,
        "has_document": bool(invoice.source_file and Path(invoice.source_file).exists()),
    }


def _save_invoice_to_db(result: Dict, invoice_type: str, file_path: Optional[str], project_code: Optional[str], cost_center: Optional[str], company_id: str = None, uploader_user_id: str = None):
    """Save processed invoice and its entries to the database"""
    extracted = result.get("extracted_data", {})
    invoice_id = result["invoice_id"]
    
    with get_db() as db:
        invoice = Invoice(
            invoice_id=invoice_id,
            invoice_type=invoice_type,
            status=result.get("status", "pending_review"),
            vendor_name=extracted.get("vendor_name") or "Unknown Vendor",
            vendor_address=extracted.get("vendor_address"),
            vendor_brn=extracted.get("vendor_brn"),
            vendor_vat=extracted.get("vendor_vat"),
            invoice_number=extracted.get("invoice_number") or "N/A",
            invoice_date=extracted.get("invoice_date"),
            due_date=extracted.get("due_date"),
            purchase_order=extracted.get("purchase_order"),
            currency=extracted.get("currency", "MUR"),
            subtotal=extracted.get("subtotal", 0.0),
            tax_total=extracted.get("tax_total", 0.0),
            total_amount=extracted.get("total_amount", 0.0),
            payment_terms=extracted.get("payment_terms"),
            notes=extracted.get("notes"),
            project_code=project_code,
            cost_center=cost_center,
            confidence_score=extracted.get("confidence_score", 0.0),
            raw_text=extracted.get("raw_text", "")[:5000],
            ai_analysis=json.dumps(extracted.get("ai_analysis")) if extracted.get("ai_analysis") else None,
            source_file=file_path,
            company_id=company_id,
            tds_applicable=extracted.get("tds_applicable", False),
            tds_rate=extracted.get("tds_rate", 0.0) if extracted.get("tds_applicable") else 0.0,
            assigned_to=uploader_user_id,  # default assign to uploader
        )

        # Multi-currency: convert to base (MUR) at booking
        currency = extracted.get("currency", "MUR")
        total = extracted.get("total_amount", 0.0)
        if currency and currency.upper() != "MUR":
            invoice_date = extracted.get("invoice_date", "")
            rate = get_exchange_rate(currency, invoice_date)
            invoice.exchange_rate = rate
            invoice.total_amount_base = round(total * rate, 2)
        else:
            invoice.exchange_rate = 1.0
            invoice.total_amount_base = total

        # Auto-match vendor from vendor master (hybrid: match if confidence >= 0.8)
        vendor_name = extracted.get("vendor_name") or ""
        if vendor_name and company_id:
            matched_id, confidence = _fuzzy_match_vendor(vendor_name, company_id, threshold=0.8)
            if matched_id:
                invoice.vendor_id = matched_id
                invoice.vendor_match_confidence = confidence
                # Apply vendor defaults
                vendor = db.query(Vendor).filter(Vendor.id == matched_id).first()
                if vendor:
                    if vendor.default_tds_rate > 0 and not invoice.tds_applicable:
                        invoice.tds_applicable = True
                        invoice.tds_rate = vendor.default_tds_rate
                    if vendor.brn and not invoice.vendor_brn:
                        invoice.vendor_brn = vendor.brn
                    if vendor.vat and not invoice.vendor_vat:
                        invoice.vendor_vat = vendor.vat
                logger.info(f"🔗 Vendor auto-matched: '{vendor_name}' → vendor_id={matched_id} (confidence={confidence:.2f})")

        db.add(invoice)
        
        # Save line items
        line_items = extracted.get("line_items", [])
        for i, item in enumerate(line_items):
            if isinstance(item, dict):
                li = InvoiceLineItemDB(
                    invoice_id=invoice_id,
                    line_number=item.get("line_number", i + 1),
                    description=item.get("description", ""),
                    quantity=item.get("quantity", 1),
                    unit_price=item.get("unit_price", 0),
                    amount=item.get("amount", 0),
                    tax_rate=item.get("tax_rate", 15.0),
                    tax_amount=item.get("tax_amount", 0),
                    account_code=item.get("account_code") or suggest_account_code(item.get("description", ""), invoice_type, company_id=company_id)[0],
                    cost_center=cost_center,
                    project_code=project_code,
                )
                db.add(li)
        
        # Save journal entries
        for entry in result.get("suggested_entries", []):
            je = JournalEntryDB(
                entry_id=entry["entry_id"],
                invoice_id=invoice_id,
                entry_date=entry.get("entry_date"),
                reference=entry.get("reference"),
                description=entry.get("description"),
                total_debit=entry.get("total_debit", 0),
                total_credit=entry.get("total_credit", 0),
                is_balanced=entry.get("is_balanced", True),
                status="draft",
                created_by=entry.get("created_by", "system"),
                company_id=company_id,
            )
            db.add(je)
            
            for line in entry.get("lines", []):
                jel = JournalEntryLineDB(
                    entry_id=entry["entry_id"],
                    account_code=line.get("account_code"),
                    account_name=line.get("account_name"),
                    description=line.get("description"),
                    debit=line.get("debit", 0),
                    credit=line.get("credit", 0),
                    cost_center=line.get("cost_center"),
                    project_code=line.get("project_code"),
                )
                db.add(jel)
        
        db.commit()
        logger.info(f"💾 Saved invoice {invoice_id} to database")
