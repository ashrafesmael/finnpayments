#!/usr/bin/env python3
"""
Export Journal Entries to Excel - FinnPayments
==============================================
Adds:
1. Backend endpoint: GET /accounting/export/excel
2. Frontend API client function
3. Export button on Journal Entries page
"""

# ============================================================
# PATCH 1: api.py - Add Excel export endpoint
# ============================================================

file = "/home/administrator/finnpayments/src/api.py"
with open(file, "r") as f:
    content = f.read()

# Add FileResponse import
if "FileResponse" not in content:
    content = content.replace(
        "from fastapi.responses import JSONResponse",
        "from fastapi.responses import JSONResponse, FileResponse"
    )
    print("✅ Added FileResponse import")
else:
    print("✅ FileResponse import already exists")

# Add export endpoint after the reverse_journal_entry endpoint
export_endpoint = '''

@app.get("/accounting/export/excel")
async def export_journal_entries_excel(
    status: Optional[str] = "posted",
):
    """Export journal entries to Excel. Defaults to posted entries only."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import tempfile

    with get_db() as db:
        query = db.query(JournalEntryDB).order_by(JournalEntryDB.entry_date.asc(), JournalEntryDB.created_at.asc())
        if status:
            query = query.filter(JournalEntryDB.status == status)
        entries = query.all()

        if not entries:
            raise HTTPException(status_code=404, detail="No journal entries found to export")

        wb = openpyxl.Workbook()

        # --- Sheet 1: Journal Entries Detail ---
        ws = wb.active
        ws.title = "Journal Entries"

        # Styles
        header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1B4332")
        data_font = Font(name="Arial", size=10)
        accent_font = Font(name="Arial", size=10, color="1B7A43")
        debit_font = Font(name="Arial", size=10, color="C0392B")
        credit_font = Font(name="Arial", size=10, color="27AE60")
        total_font = Font(name="Arial", size=10, bold=True)
        total_fill = PatternFill("solid", fgColor="F0F4F0")
        border = Border(
            bottom=Side(style="thin", color="D5D5D5")
        )
        currency_fmt = '#,##0.00'

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "MC Golf - Journal Entries Export"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1B4332")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888")
        ws["A3"] = f"Filter: {status or 'All'} | Total entries: {len(entries)}"
        ws["A3"].font = Font(name="Arial", size=9, color="888888")

        # Headers
        headers = ["Entry ID", "Date", "Vendor / Reference", "Account Code", "Account Name", "Description", "Debit", "Credit"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 6
        grand_debit = 0
        grand_credit = 0

        for entry in entries:
            # Get vendor name from related invoice
            vendor = ""
            if entry.invoice:
                vendor = entry.invoice.vendor_name or ""

            entry_start_row = row
            for line in entry.lines:
                ws.cell(row=row, column=1, value=entry.entry_id).font = accent_font
                ws.cell(row=row, column=2, value=entry.entry_date).font = data_font
                ws.cell(row=row, column=3, value=vendor or entry.reference or "").font = data_font
                ws.cell(row=row, column=4, value=line.account_code).font = Font(name="Consolas", size=10)
                ws.cell(row=row, column=5, value=line.account_name).font = data_font
                ws.cell(row=row, column=6, value=line.description or "").font = data_font

                debit_cell = ws.cell(row=row, column=7, value=line.debit if line.debit > 0 else None)
                debit_cell.font = debit_font
                debit_cell.number_format = currency_fmt

                credit_cell = ws.cell(row=row, column=8, value=line.credit if line.credit > 0 else None)
                credit_cell.font = credit_font
                credit_cell.number_format = currency_fmt

                for c in range(1, 9):
                    ws.cell(row=row, column=c).border = border

                grand_debit += line.debit or 0
                grand_credit += line.credit or 0
                row += 1

            # Entry subtotal row
            ws.cell(row=row, column=6, value=f"Entry Total ({entry.entry_id})").font = total_font
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
            sub_debit = ws.cell(row=row, column=7, value=entry.total_debit)
            sub_debit.font = total_font
            sub_debit.number_format = currency_fmt
            sub_debit.fill = total_fill
            sub_credit = ws.cell(row=row, column=8, value=entry.total_credit)
            sub_credit.font = total_font
            sub_credit.number_format = currency_fmt
            sub_credit.fill = total_fill
            row += 1  # blank row separator

        # Grand total
        row += 1
        ws.cell(row=row, column=6, value="GRAND TOTAL").font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
        gt_debit = ws.cell(row=row, column=7, value=grand_debit)
        gt_debit.font = Font(name="Arial", bold=True, size=11, color="C0392B")
        gt_debit.number_format = currency_fmt
        gt_credit = ws.cell(row=row, column=8, value=grand_credit)
        gt_credit.font = Font(name="Arial", bold=True, size=11, color="27AE60")
        gt_credit.number_format = currency_fmt

        # Column widths
        widths = [22, 12, 25, 14, 25, 45, 15, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A6"

        # --- Sheet 2: Summary by Account ---
        ws2 = wb.create_sheet("Summary by Account")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = "Journal Entries Summary by Account"
        ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1B4332")

        summary_headers = ["Account Code", "Account Name", "Total Debit", "Total Credit", "Net"]
        for col, h in enumerate(summary_headers, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Aggregate by account
        account_totals = {}
        for entry in entries:
            for line in entry.lines:
                key = (line.account_code or "", line.account_name or "")
                if key not in account_totals:
                    account_totals[key] = {"debit": 0, "credit": 0}
                account_totals[key]["debit"] += line.debit or 0
                account_totals[key]["credit"] += line.credit or 0

        srow = 4
        for (code, name), totals in sorted(account_totals.items()):
            ws2.cell(row=srow, column=1, value=code).font = Font(name="Consolas", size=10)
            ws2.cell(row=srow, column=2, value=name).font = data_font
            d = ws2.cell(row=srow, column=3, value=totals["debit"])
            d.font = debit_font
            d.number_format = currency_fmt
            c = ws2.cell(row=srow, column=4, value=totals["credit"])
            c.font = credit_font
            c.number_format = currency_fmt
            net = ws2.cell(row=srow, column=5, value=totals["debit"] - totals["credit"])
            net.font = data_font
            net.number_format = currency_fmt
            for col in range(1, 6):
                ws2.cell(row=srow, column=col).border = border
            srow += 1

        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 15
        ws2.column_dimensions["E"].width = 15
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A4"

        # Save to temp file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir="temp_uploads")
        wb.save(tmp.name)
        tmp.close()

        filename = f"MC_Golf_Journal_Entries_{status or 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
'''

if "export_journal_entries_excel" in content:
    print("✅ Export endpoint already exists in api.py")
else:
    # Insert before the suggest_account endpoint or at the end of accounting section
    insert_marker = "@app.get(\"/accounting/suggest-account\")"
    if insert_marker in content:
        content = content.replace(insert_marker, export_endpoint + "\n" + insert_marker)
    else:
        # Try another insertion point
        insert_marker2 = "@app.get(\"/accounting/chart-of-accounts\")"
        if insert_marker2 in content:
            content = content.replace(insert_marker2, export_endpoint + "\n" + insert_marker2)
        else:
            # Append before _save_invoice_to_db
            content = content.replace("def _save_invoice_to_db", export_endpoint + "\n\ndef _save_invoice_to_db")
    print("✅ Added export_journal_entries_excel endpoint")

with open(file, "w") as f:
    f.write(content)
print(f"✅ Saved {file}")


# ============================================================
# PATCH 2: Frontend api.js - Add export function
# ============================================================

file2 = "/home/administrator/finnpayments/frontend/src/services/api.js"
with open(file2, "r") as f:
    content2 = f.read()

export_fn = '''

// ─── Export ──────────────────────────────────────────────
export const exportJournalEntriesExcel = async (status = 'posted') => {
  const url = `${API_BASE}/accounting/export/excel?status=${status}`;
  const response = await fetch(url);
  if (!response.ok) {
    const error = await response.json().catch(() => ({ detail: response.statusText }));
    throw new Error(error.detail || `Export failed: ${response.status}`);
  }
  const blob = await response.blob();
  const downloadUrl = window.URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = downloadUrl;
  a.download = response.headers.get('content-disposition')?.split('filename=')[1]?.replace(/"/g, '') || `journal_entries_${status}.xlsx`;
  document.body.appendChild(a);
  a.click();
  a.remove();
  window.URL.revokeObjectURL(downloadUrl);
};
'''

if "exportJournalEntriesExcel" in content2:
    print("✅ Export function already exists in api.js")
else:
    # Add before the health check section
    content2 = content2.replace(
        "// ─── Health",
        export_fn + "// ─── Health"
    )
    print("✅ Added exportJournalEntriesExcel to api.js")

with open(file2, "w") as f:
    f.write(content2)
print(f"✅ Saved {file2}")


# ============================================================
# PATCH 3: Frontend App.jsx - Add import + export button
# ============================================================

file3 = "/home/administrator/finnpayments/frontend/src/App.jsx"
with open(file3, "r") as f:
    content3 = f.read()

# 3a. Add import
if "exportJournalEntriesExcel" not in content3:
    content3 = content3.replace(
        "postJournalEntry, reverseJournalEntry, getChartOfAccounts, checkHealth",
        "postJournalEntry, reverseJournalEntry, getChartOfAccounts, checkHealth,\n  exportJournalEntriesExcel"
    )
    print("✅ Added exportJournalEntriesExcel import")
else:
    print("✅ Import already exists")

# 3b. Add Download icon
if "Download:" not in content3:
    content3 = content3.replace(
        "  Upload: () =>",
        '  Download: () => <svg width="18" height="18" fill="none" stroke="currentColor" strokeWidth="1.5" viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7,10 12,15 17,10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>,\n  Upload: () =>'
    )
    print("✅ Added Download icon")

# 3c. Add export button to Journal Entries page header
old_je_header = '<div className="page-header"><h2>Journal Entries</h2></div>'
new_je_header = '''<div className="page-header">
        <h2>Journal Entries</h2>
        <button className="btn btn-primary" onClick={async () => {
          try { await exportJournalEntriesExcel(filter || 'posted'); }
          catch (e) { alert(e.message); }
        }}><Icons.Download /> Export to Excel</button>
      </div>'''

if "Export to Excel" in content3:
    print("✅ Export button already exists")
else:
    content3 = content3.replace(old_je_header, new_je_header)
    print("✅ Added Export to Excel button")

with open(file3, "w") as f:
    f.write(content3)
print(f"✅ Saved {file3}")


print("\n" + "="*60)
print("JOURNAL ENTRIES EXCEL EXPORT - DEPLOYMENT COMPLETE")
print("="*60)
print("""
Features:
  - GET /accounting/export/excel?status=posted
  - Sheet 1: "Journal Entries" - full detail with entry ID, date, vendor,
    account code, account name, description, debit, credit
  - Sheet 2: "Summary by Account" - aggregated totals per GL account
  - Professional formatting: MC Golf green theme, frozen headers
  - Frontend: "Export to Excel" button on Journal Entries page
  - Respects current filter (draft/posted/all)

Restart to apply:
  sudo fuser -k 3001/tcp 8001/tcp 2>/dev/null; sleep 2
  cd ~/finnpayments && ./start-all.sh
""")
